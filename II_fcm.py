# -*- coding: utf-8 -*-
"""
Created on Wed Mar 27 15:51:29 2024

@author: Jonas Schuster
"""

# =============================================================================
# Define program
# =============================================================================
program_type = input("Please select the program type:\n1. Manually, one sample\n2. Automatically, many samples with subfolders\n3. Automatically, many samples without subfolder\n************************\n")

# Convert the input to an integer
program_type = int(program_type)

# Perform actions based on the selected program type
if program_type == 1:
    print("************************\nYou selected 1. program.\n************************\n")
    dil_fac = input("Please select the dilution factor: \n1, 10 or 100 if measured offline \n2 if measured online via OC300\n************************\n")
    dil_fac = float(dil_fac)
    
elif program_type == 2:
    print("************************\nYou selected 2. program.\n************************")
    dil_fac = input("Please select the dilution factor: \n1, 10 or 100 if measured offline \n2 if measured online via OC300\n************************\n")
    dil_fac = float(dil_fac)
    sample_id = input("************************\nPlease select how many different samples are measured: 1, 2 or 3?\n************************\n")
    sample_id = int(sample_id)
    sample_id_start = input("************************\nSample ID of starting analyzis: 1, 2 or 3?\n************************\n")
    sample_id_start = int(sample_id_start)
        
elif program_type == 3:
    print("************************\nYou selected 3. program.\n************************\n")
    dil_fac = input("Please select the dilution factor: \n1, 10 or 100 if measured offline \n2 if measured online via OC300\n************************\n")
    dil_fac = float(dil_fac)
    sample_id = input("************************\nPlease select how many different samples are measured: 1, 2 or 3?\n************************\n")
    sample_id = int(sample_id)
    sample_id_start = input("************************\nSample ID of starting analyzis: 1, 2 or 3?\n************************\n")
    sample_id_start = int(sample_id_start)
    
else:
    print("Invalid program type selected.")
       
# =============================================================================
# Perform actions based on the selected program type
# =============================================================================
if program_type == 1:
    Dictionary = input( 'Please enter the ADDRESS and NAME of the fcs.-file and press Enter. Example: "C:/Users/Public/flow_cytometry/raw_data_example/ABC.fcs"\n************************\n ')
    print("************************************************")
    
elif program_type == 2:
    Dictionary = input( 'Please enter the ADDRESS of the folder containing subfolders and press Enter. Example: "C:/Users/Public/flow_cytometry/raw_data/"\n************************\n ')
    print("************************************************")
    
elif program_type == 3:
    Dictionary = input( 'Please enter the ADDRESS of the folder containing files and press Enter. Example: "C:/Users/Public/flow_cytometry/raw_data/"\n************************\n')
    print("************************************************")

# In case of an automatic data analysis. The time entered here should be the time difference between two measurements
if program_type == 2:
    t = int(input('Please enter the time (in seconds) between two measurements. (e.g. 900 equals 15 min)\n'))
    print("************************************************")
    
elif program_type == 3:
    t = int(input('Please enter the time (in seconds) between two measurements. (e.g. 900 equals 15 min)\n'))
    print("************************************************")
    
#Location and name for results
location = input('Enter the location of the folder for the excel sheet containing the final results. Example: "C:/Users/Public/flow_cytometry/"\n')
print("************************************************")

filename = input("Enter the filename for the Excel file ('fcm_results' for online analysis'):\n") + ".xlsx"
print("************************************************")

confirmation = input("\nWAIT here and press enter when the first flow cytometry measurement is completed and the raw data is stored in the folder\n")


# =============================================================================
# Data analysis
# =============================================================================
import bokeh
from bokeh.plotting import show
import matplotlib.pyplot as plt
import flowkit as fk
import time
import numpy as np
from itertools import cycle, islice
import csv
import matplotlib.pyplot as plt
from matplotlib.colors import LinearSegmentedColormap
from matplotlib.cm import get_cmap
import mpl_scatter_density
import os
import pandas as pd
from tabulate import tabulate
from openpyxl import load_workbook
bokeh.io.output_notebook()
%matplotlib inline
_ = plt.ioff()

# check version so users can verify they have the same version\API
fk.__version__

# =============================================================================
#   Define gate boundaries
# =============================================================================
    # TCC gate
x2 = 50
x4 = 10000
x5 = 50000
y1 = 10
y2 = 60
y5 = 10000

# HNA/LNA boundary on x-axis, value is minimum in histogram between peak regions for mineral water (Evian)
x_LNA = 315.0

# lienar diagonal equation TCC gate
m = (y5 - y2)/(x4 - x2)
b = y2 - (m * x2)

# y-upper HNA/LNA boundary
y_LNA = m * x_LNA + b

# define the points for TCC gate
point5 = (x5, y1)
point1 = (x2, y1)
point4 = (x5, y5)
point2 = (x2, y2)
point3 = (x4, y5)

# define the points for LNA gate
point6 = (x_LNA, y1)
point7 = (x_LNA, y_LNA)
    
# create TCC Gate line
x_line = [point1[0], point2[0], point3[0], point4[0], point5[0], point1[0]]
y_line = [point1[1], point2[1], point3[1], point4[1], point5[1], point1[1]]

# create HNA/LNA boundary
x_line_LNA = [point6[0], point7[0]]
y_line_LNA = [point6[1], point7[1]]


# =============================================================================
# 1. manual program
# =============================================================================
if program_type == 1:
    
# =============================================================================
#   Read raw data and extract general information
# =============================================================================
    # Load FCS file using FlowKit
    sample = fk.Sample(Dictionary)

    # Print the sample information
    print(sample)
    
    # Get string representation of sample
    sample_string = str(sample)

    # Extract name from sample string
    name = sample_string.split(',')[1].strip()
    
    # Get metadata
    metadata=sample.get_metadata()
    
    # Get acquisition volume, volume in metadata is stored in nL, conversion to mL
    acq_vol = float(metadata['vol'])/1e6

    # Print acquisition volume
    print(f'Volume of the sample is: {acq_vol} mL')
        
# =============================================================================
#   Extract fluorescence detectors information and count
# =============================================================================
    # extract array with all data needed, TCC
    fcs_raw = sample._orig_events

    # raw data in an array
    fcs_raw = sample.as_dataframe(source='raw')
    # remove second header from the array
    fcs_raw.columns = fcs_raw.columns.droplevel(level=1)

    # determine array columns for CyStain Green and CyStain Red fluorescence detectors, TCC
    tcc_green = fcs_raw.iloc[:,[2]]
    tcc_green = tcc_green['CyStain Green'].to_numpy()
    tcc_green = tcc_green.astype(int)
    tcc_red = fcs_raw.iloc[:,[4]]
    tcc_red = tcc_red['CyStain Red'].to_numpy()
    tcc_red = tcc_red.astype(int)
    
    # count within the gates
    # create dataframe for total cell count
    df_TCC = pd.DataFrame({'CyStain Green': tcc_green, 'CyStain Red': tcc_red})

    # define cell count for TCC plot
    # define rectangle base
    # x boundaries
    df1 = df_TCC[df_TCC['CyStain Green']>x2]
    df1 = df1[df1['CyStain Green']<x5]
    # y boundaries
    df1 = df1[df1['CyStain Red']>y1]
    df1 = df1[df1['CyStain Red']<y5].reset_index(drop=True)

    # define the length of the rectangle polygon TCC gate
    length_df1 = len(df1.index)

    # create new and empty dataframe for counts in polygon with diagonal slope
    df_TCC_new = pd.DataFrame(columns=['CyStain Green','CyStain Red'])

    for i in range(length_df1):
        if (m*df1.iloc[i,0]+b) >= df1.iloc[i,1]:
            df_TCC_new = pd.concat([df_TCC_new, df1.iloc[[i],:]], ignore_index=True)
    
    # the number of rows of df_TCC_new is the number of counts within the Gate
    cell_number_gate = len(df_TCC_new)

    # target volume for count in mL
    v = 1

    # factor for the conversion in cells per mL
    volume_factor = v/acq_vol

    # cell count per mL (TCC), includes conversion in mL and manual dilution factor
    TCC_per_mL = int(volume_factor * cell_number_gate) * dil_fac

    # calculate HNA/LNA cells/mL
    # HNA/LNA boundary on x-axis, value is minimum in histogram between peak regions for mineral water (Evian)
    x_HNA = x_LNA

    # create array for counting
    tcc_green_new = df_TCC_new.iloc[:,[0]].to_numpy()
    
    # create empty lists for HNA and LNA counts
    list_HNA = []
    list_LNA = []
    
    # fill the lists
    for l in range(len(tcc_green_new)):
        if tcc_green_new[l] > x_HNA:
            list_HNA.append(l)
        else:
            list_LNA.append(l)
            
    # cell count per mL (HNA and LNA), includes conversion in mL and manual dilution factor
    HNA_per_mL = int(volume_factor * len(list_HNA)) * dil_fac
    LNA_per_mL = int(volume_factor * len(list_LNA)) * dil_fac
    HNA_amount = HNA_per_mL / TCC_per_mL *100

    print('\nTotal cells per mL (offline):')
    print(TCC_per_mL)
    print('\nHNA cells per mL (offline):')
    print(HNA_per_mL)
    print('\nLNA cells per mL (offline):')
    print(LNA_per_mL)
    print('\nAmount of HNA cells (offline):')
    print(HNA_amount)
      
# =============================================================================
#   Plot of the fcm fingerprint
# =============================================================================
    # define font type
    plt.rcParams['font.family']='Times New Roman'
    
    # define the font style for tick notation
    def format_y_tick(x, pos):
        """Format y-axis tick labels in x10 notation."""
        if x == 0:
            return '0'
        exponent=int(np.log10(abs(x)))
        coefficient=x / 10 ** exponent
        if coefficient == 1:
            # added font family setting
            return r'$\mathregular{{10^{{\mathregular{{{}}}}}}}$'.format(exponent)
        elif coefficient.is_integer():
            # added font family setting
            return r'$\mathregular{{{:d}\times10^{{{}}}}}$'.format(int(coefficient), exponent)
        else:
            # added font family setting
            return r'$\mathregular{{{:0.1f}\times10^{{{}}}}}$'.format(coefficient, exponent)

    # chose colormap 
    magma_white_background = LinearSegmentedColormap.from_list('magma_white_background', [
        (0, '#ffffff'),    # White
        (1e-20, '#000003'), # Dark brown
        (0.2, '#530002'),   # Brown
        (0.4, '#bd2200'),   # Orange
        (0.6, '#fcaa00'),   # Yellow
        (0.8, '#fcf600'),   # Light yellow
        (1, '#ffffff')      # White
    ], N=256)

    # create the fcm fingerprint
    fig, ax = plt.subplots(1, 1, figsize=(8, 6), subplot_kw={'projection': 'scatter_density'})
    fig.subplots_adjust(left=0.15, right=0.95, bottom=0.15, top=0.95)
    density = ax.scatter_density(tcc_green, tcc_red, cmap=magma_white_background, vmin=0, vmax=10)
    # Colorbar with adjusted fontsize
    cbar = plt.colorbar(density, ax=ax)
    cbar.set_label('Number of points per pixel', fontsize=20, labelpad=10)  # Adjust the fontsize as needed
    cbar.ax.tick_params(axis='both', which='major', labelsize=20)
    # general settings
    ax.set_xlim(10,60000)
    ax.set_ylim(8,50000)
    ax = plt.yscale('log')
    ax = plt.xscale('log')
    ax = plt.xlabel('Green fluorescence in A.U.', fontsize=20, labelpad=10)
    ax = plt.ylabel('Red fluorescence in A.U.', fontsize=20, labelpad=10)
    ax = plt.title(name)
    ax = plt.xticks(fontsize=20)
    ax = plt.yticks(fontsize=20)
    # plot TCC gate
    ax = plt.plot(x_line,y_line, color='black')
    # plot LNA gate
    ax = plt.plot(x_line_LNA, y_line_LNA, color='black')    
    ax = plt.text(70, 11, 'LNA', fontsize=20, color='black')
    # plot HNA gate
    ax = plt.text(1000, 11, 'HNA', fontsize=20, color='black')
    plt.show()
    
# =============================================================================
#   Export results
# =============================================================================
    # Create a pandas DataFrame containing the results
    df = pd.DataFrame({"Parameter": ["LNA", "HNA", "TCC", "HNA in %"],
                        "Counts [per mL]": [LNA_per_mL, HNA_per_mL, TCC_per_mL, HNA_amount],
                        "Counts [abs.]": [len(list_LNA), len(list_LNA), len(df_TCC_new), HNA_amount]})

    from datetime import datetime
    import openpyxl

    # Prompt the user to enter the file name and location
    file_name = filename
    file_path = location

    # Load the existing Excel workbook
    try:
        workbook = openpyxl.load_workbook(file_path + file_name)
    except FileNotFoundError:
        workbook = openpyxl.Workbook()

    # Set the active worksheet
    worksheet = workbook.active

    # If the worksheet doesn't already have headers, add them
    if worksheet.cell(row=1, column=1).value is None:
        worksheet.cell(row=1, column=1, value="Index")
        worksheet.cell(row=1, column=2, value="File Name")
        worksheet.cell(row=1, column=3, value="HNA")
        worksheet.cell(row=1, column=4, value="LNA")
        worksheet.cell(row=1, column=5, value="TCC")
        worksheet.cell(row=1, column=6, value="HNA in %")
        worksheet.cell(row=1, column=7, value="Time")
        worksheet.cell(row=1, column=8, value="Date")

    # Write the results to the Excel worksheet
    now = datetime.now()
    date_string = now.strftime("%Y-%m-%d")
    time_string = now.strftime("%H:%M:%S")
    filename_parts = name.split("__")
    index = filename_parts[-1].split(".")[0]

    # Write the results to the Excel worksheet
    worksheet.append([index, name, HNA_per_mL, LNA_per_mL, TCC_per_mL, HNA_amount, time_string, date_string])

    # Adjust the width of the "File Name" column
    max_name_length = len(name)
    if max_name_length > 30:
        max_name_length = 30
    worksheet.column_dimensions['B'].width = max_name_length + 2

    # Adjust the width of the "File Name" column
    max_name_length_date_string = len(date_string)
    if max_name_length_date_string > 30:
        max_name_length_date_string = 30
    worksheet.column_dimensions['H'].width = max_name_length_date_string + 2

    # Save the Excel workbook
    workbook.save(file_path + file_name)

    # Print the table using tabulate
    print(tabulate(df, headers="keys", tablefmt="psql"))

    print("*********************************************************************************************************")

# =============================================================================
# 2. Automatics with subfolder program  
# =============================================================================
elif program_type == 2:
    
# =============================================================================
#     print('Waiting for first sample...')
# 
#     # Wait for t seconds before checking for a new file
#     time.sleep(t)
# =============================================================================
    while True:
        
# =============================================================================
#       Check subfolder structure and find the latest sample
# =============================================================================
        # Set the directory path where the FCS files are located
        parent_directory_path = Dictionary 

        # Get the list of all subdirectories and files within the parent directory recursively
        all_contents = []
        for root, dirs, files in os.walk(parent_directory_path):
            all_contents.extend(dirs)
            all_contents.extend(files)

        # Filter the list to only include directories
        subdirectories = [subdir for subdir in all_contents if os.path.isdir(os.path.join(parent_directory_path, subdir))]

        # Find the newest subdirectory based on its creation time
        newest_subdirectory = max(subdirectories, key=lambda x: os.path.getctime(os.path.join(parent_directory_path, x)))

        # Construct the path to the newest subdirectory
        subdirectory_path = os.path.join(parent_directory_path, newest_subdirectory)

        # Get the list of files in the subdirectory
        file_list = os.listdir(subdirectory_path)

        # Filter the list to only include FCS files
        fcs_files = [file for file in file_list if file.endswith('.fcs')]

        if not fcs_files:
            print("No FCS files found in the newest subfolder.")

            print("*********************************************************************************************************")

            # Wait for t/100 seconds before checking for a new file
            time.sleep(t/100)

        else: 
            # Find the newest FCS file based on its creation time
            newest_file = max(fcs_files, key=lambda x: os.path.getctime(os.path.join(subdirectory_path, x)))

            # Construct the path to the newest FCS file
            newest_file_path = os.path.join(subdirectory_path, newest_file)
            print(newest_file_path)
            
            # =============================================================================
            #   Read raw data and extract general information
            # =============================================================================
            # Load FCS file using FlowKit
            sample = fk.Sample(newest_file_path)

            # Print the sample information
            print(sample)
            
            # Get string representation of sample
            sample_string = str(sample)

            # Extract name from sample string
            name = sample_string.split(',')[1].strip()
            
            # Get metadata
            metadata=sample.get_metadata()
            
            # Get acquisition volume, volume in metadata is stored in nL, conversion to mL
            acq_vol = float(metadata['vol'])/1e6

            # Print acquisition volume
            print(f'Volume of the sample is: {acq_vol} mL')
                    
            # =============================================================================
            #   Extract fluorescence detectors information and count
            # =============================================================================
            # extract array with all data needed, TCC
            fcs_raw = sample._orig_events

            # raw data in an array
            fcs_raw = sample.as_dataframe(source='raw')
            # remove second header from the array
            fcs_raw.columns = fcs_raw.columns.droplevel(level=1)

            # determine array columns for CyStain Green and CyStain Red fluorescence detectors, TCC
            tcc_green = fcs_raw.iloc[:,[2]]
            tcc_green = tcc_green['CyStain Green'].to_numpy()
            tcc_green = tcc_green.astype(int)
            tcc_red = fcs_raw.iloc[:,[4]]
            tcc_red = tcc_red['CyStain Red'].to_numpy()
            tcc_red = tcc_red.astype(int)
            
            # count within the gates
            # create dataframe for total cell count
            df_TCC = pd.DataFrame({'CyStain Green': tcc_green, 'CyStain Red': tcc_red})

            # define cell count for TCC plot
            # define rectangle base
            # x boundaries
            df1 = df_TCC[df_TCC['CyStain Green']>x2]
            df1 = df1[df1['CyStain Green']<x5]
            # y boundaries
            df1 = df1[df1['CyStain Red']>y1]
            df1 = df1[df1['CyStain Red']<y5].reset_index(drop=True)

            # define the length of the rectangle polygon TCC gate
            length_df1 = len(df1.index)

            # create new and empty dataframe for counts in polygon with diagonal slope
            df_TCC_new = pd.DataFrame(columns=['CyStain Green','CyStain Red'])

            for i in range(length_df1):
                if (m*df1.iloc[i,0]+b) >= df1.iloc[i,1]:
                    df_TCC_new = pd.concat([df_TCC_new, df1.iloc[[i],:]], ignore_index=True)
            
            # the number of rows of df_TCC_new is the number of counts within the Gate
            cell_number_gate = len(df_TCC_new)

            # target volume for count in mL
            v = 1

            # factor for the conversion in cells per mL
            volume_factor = v/acq_vol

            # cell count per mL (TCC), includes conversion in mL and manual dilution factor
            TCC_per_mL = int(volume_factor * cell_number_gate) * dil_fac

            # calculate HNA/LNA cells/mL
            # HNA/LNA boundary on x-axis, value is minimum in histogram between peak regions for mineral water (Evian)
            x_HNA = x_LNA

            # create array for counting
            tcc_green_new = df_TCC_new.iloc[:,[0]].to_numpy()
            
            # create empty lists for HNA and LNA counts
            list_HNA = []
            list_LNA = []
            
            # fill the lists
            for l in range(len(tcc_green_new)):
                if tcc_green_new[l] > x_HNA:
                    list_HNA.append(l)
                else:
                    list_LNA.append(l)
                    
            # cell count per mL (HNA and LNA), includes conversion in mL and manual dilution factor
            HNA_per_mL = int(volume_factor * len(list_HNA)) * dil_fac
            LNA_per_mL = int(volume_factor * len(list_LNA)) * dil_fac
            HNA_amount = HNA_per_mL / TCC_per_mL * 100

            print('\nTotal cells per mL:')
            print(TCC_per_mL)
            print('\nHNA cells per mL:')
            print(HNA_per_mL)
            print('\nLNA cells per mL:')
            print(LNA_per_mL)
            print('\nAmount of HNA cells in %')
            print(HNA_amount)
                  
            # =============================================================================
            #   Plot of the fcm fingerprint
            # =============================================================================
            # define font type
            plt.rcParams['font.family']='Times New Roman'
            
            # define the font style for tick notation
            def format_y_tick(x, pos):
                """Format y-axis tick labels in x10 notation."""
                if x == 0:
                    return '0'
                exponent=int(np.log10(abs(x)))
                coefficient=x / 10 ** exponent
                if coefficient == 1:
                    # added font family setting
                    return r'$\mathregular{{10^{{\mathregular{{{}}}}}}}$'.format(exponent)
                elif coefficient.is_integer():
                    # added font family setting
                    return r'$\mathregular{{{:d}\times10^{{{}}}}}$'.format(int(coefficient), exponent)
                else:
                    # added font family setting
                    return r'$\mathregular{{{:0.1f}\times10^{{{}}}}}$'.format(coefficient, exponent)

            # chose colormap 
            magma_white_background = LinearSegmentedColormap.from_list('magma_white_background', [
                (0, '#ffffff'),    # White
                (1e-20, '#000003'), # Dark brown
                (0.2, '#530002'),   # Brown
                (0.4, '#bd2200'),   # Orange
                (0.6, '#fcaa00'),   # Yellow
                (0.8, '#fcf600'),   # Light yellow
                (1, '#ffffff')      # White
            ], N=256)

            # create the fcm fingerprint
            fig, ax = plt.subplots(1, 1, figsize=(8, 6), subplot_kw={'projection': 'scatter_density'})
            fig.subplots_adjust(left=0.15, right=0.95, bottom=0.15, top=0.95)
            density = ax.scatter_density(tcc_green, tcc_red, cmap=magma_white_background, vmin=0, vmax=10)
            # Colorbar with adjusted fontsize
            cbar = plt.colorbar(density, ax=ax)
            cbar.set_label('Number of points per pixel', fontsize=20, labelpad=10)  # Adjust the fontsize as needed
            cbar.ax.tick_params(axis='both', which='major', labelsize=20)
            # general settings
            ax.set_xlim(10,60000)
            ax.set_ylim(8,50000)
            ax = plt.yscale('log')
            ax = plt.xscale('log')
            ax = plt.xlabel('Green fluorescence in A.U.', fontsize=20, labelpad=10)
            ax = plt.ylabel('Red fluorescence in A.U.', fontsize=20, labelpad=10)
            ax = plt.title(name)
            ax = plt.xticks(fontsize=20)
            ax = plt.yticks(fontsize=20)
            # plot TCC gate
            ax = plt.plot(x_line,y_line, color='black')
            # plot LNA gate
            ax = plt.plot(x_line_LNA, y_line_LNA, color='black')    
            ax = plt.text(70, 11, 'LNA', fontsize=20, color='black')
            # plot HNA gate
            ax = plt.text(1000, 11, 'HNA', fontsize=20, color='black')
            plt.show()
                
            # =============================================================================
            #   Export results
            # =============================================================================
            # Create a pandas DataFrame containing the results
            df = pd.DataFrame({"Parameter": ["LNA", "HNA", "TCC", "HNA in %"],
                                "Counts [per mL]": [LNA_per_mL, HNA_per_mL, TCC_per_mL, HNA_amount],
                                "Counts [abs.]": [len(list_LNA), len(list_LNA), len(df_TCC_new), HNA_amount]})

            from datetime import datetime
            import openpyxl

            # Prompt the user to enter the file name and location
            file_name = filename
            file_path = location

            # Load the existing Excel workbook
            try:
                workbook = openpyxl.load_workbook(file_path + file_name)
            except FileNotFoundError:
                workbook = openpyxl.Workbook()

            # Set the active worksheet
            worksheet = workbook.active

            # If the worksheet doesn't already have headers, add them
            if worksheet.cell(row=1, column=1).value is None:
                worksheet.cell(row=1, column=1, value="Index")
                worksheet.cell(row=1, column=2, value="File Name")
                worksheet.cell(row=1, column=3, value="HNA")
                worksheet.cell(row=1, column=4, value="LNA")
                worksheet.cell(row=1, column=5, value="TCC")
                worksheet.cell(row=1, column=6, value="HNA in %")
                worksheet.cell(row=1, column=7, value="Time")
                worksheet.cell(row=1, column=8, value="Date")
                worksheet.cell(row=1, column=9, value='Sample ID')

            # Write the results to the Excel worksheet
            now = datetime.now()
            date_string = now.strftime("%Y-%m-%d")
            time_string = now.strftime("%H:%M:%S")
            filename_parts = name.split("__")
            index = filename_parts[-1].split(".")[0]
            
            number_rows_samples = worksheet.max_row
            
            #create a series with only sample id's to safe the rigth sample name in final excel
            def create_repeating_series(length, start_number=1):

                index_mapping = {1: 0, 2: 1, 3: 2}
                start_index = index_mapping[start_number]

                if sample_id == 3:
                    repeating_sequence = cycle(range(1, 4))
                    repeating_list = list(islice(repeating_sequence, start_index, start_index + length))
                elif sample_id == 2:
                    repeating_sequence = cycle(range(1, 3))
                    repeating_list = list(islice(repeating_sequence, start_index, start_index + length))
                elif sample_id == 1:
                    repeating_list = sample_id_start
                    #repeating_sequence = cycle(range(1, 2))
                    #repeating_list = list(islice(repeating_sequence, start_index, start_index + length))
                return pd.Series(repeating_list, dtype=int)

            # Define the desired length of the series and the starting number
            length_parameter = number_rows_samples
            start_number_parameter = sample_id_start

            # Create the pandas Series
            result_series = create_repeating_series(length_parameter, start_number_parameter)
            
            # define the current sample bnumber
            sample_id_current = int(result_series.iloc[-1]) 

            # Write the results to the Excel worksheet
            worksheet.append([index, name, HNA_per_mL, LNA_per_mL, TCC_per_mL, HNA_amount, time_string, date_string, sample_id_current])

            # Adjust the width of the "File Name" column
            max_name_length = len(name)
            if max_name_length > 30:
                max_name_length = 30
            worksheet.column_dimensions['B'].width = max_name_length + 2

            # Adjust the width of the "File Name" column
            max_name_length_date_string = len(date_string)
            if max_name_length_date_string > 30:
                max_name_length_date_string = 30
            worksheet.column_dimensions['H'].width = max_name_length_date_string + 2

            # Save the Excel workbook
            workbook.save(file_path + file_name)

            # Print the table using tabulate
            print(tabulate(df, headers="keys", tablefmt="psql"))

            print("*********************************************************************************************************")          
            print('Waiting for measurement')
            print("*********************************************************************************************************")
            
            # Wait for t seconds before checking for a new file
            time.sleep(t)           
# =============================================================================
# 3. Automatics without subfolder program
# =============================================================================
elif program_type == 3:
    
    # Wait for t seconds before checking for a new file
    #time.sleep(t) 
    # Set the directory path where the FCS files are located
    directory_path = Dictionary 
   
    while True:
        
        # Get the list of files in the directory
        file_list = os.listdir(directory_path)

        # Filter the list to only include FCS files
        fcs_files = [file for file in file_list if file.endswith('.fcs')]

        # Find the newest FCS file based on its creation time
        newest_file = max(fcs_files, key=lambda x: os.path.getctime(directory_path + x))

        # Construct the path to the newest FCS file
        newest_file_path = directory_path + newest_file

        print(newest_file_path)

        # =============================================================================
        #   Read raw data and extract general information
        # =============================================================================
        # Load FCS file using FlowKit
        sample = fk.Sample(newest_file_path)
        
        # Print the sample information
        print(sample)
        
        # Get string representation of sample
        sample_string = str(sample)
        
        # Extract name from sample string
        name = sample_string.split(',')[1].strip()
        
        # Get metadata
        metadata=sample.get_metadata()
        
        # Get acquisition volume, volume in metadata is stored in nL, conversion to mL
        acq_vol = float(metadata['vol'])/1e6
        
        # Print acquisition volume
        print(f'Volume of the sample is: {acq_vol} mL')
                
        # =============================================================================
        #   Extract fluorescence detectors information and count
        # =============================================================================
        # extract array with all data needed, TCC
        fcs_raw = sample._orig_events
        
        # raw data in an array
        fcs_raw = sample.as_dataframe(source='raw')
        # remove second header from the array
        fcs_raw.columns = fcs_raw.columns.droplevel(level=1)
        
        # determine array columns for CyStain Green and CyStain Red fluorescence detectors, TCC
        tcc_green = fcs_raw.iloc[:,[2]]
        tcc_green = tcc_green['CyStain Green'].to_numpy()
        tcc_green = tcc_green.astype(int)
        tcc_red = fcs_raw.iloc[:,[4]]
        tcc_red = tcc_red['CyStain Red'].to_numpy()
        tcc_red = tcc_red.astype(int)
        
        # count within the gates
        # create dataframe for total cell count
        df_TCC = pd.DataFrame({'CyStain Green': tcc_green, 'CyStain Red': tcc_red})
        
        # define cell count for TCC plot
        # define rectangle base
        # x boundaries
        df1 = df_TCC[df_TCC['CyStain Green']>x2]
        df1 = df1[df1['CyStain Green']<x5]
        # y boundaries
        df1 = df1[df1['CyStain Red']>y1]
        df1 = df1[df1['CyStain Red']<y5].reset_index(drop=True)
        
        # define the length of the rectangle polygon TCC gate
        length_df1 = len(df1.index)
        
        # create new and empty dataframe for counts in polygon with diagonal slope
        df_TCC_new = pd.DataFrame(columns=['CyStain Green','CyStain Red'])
        
        for i in range(length_df1):
            if (m*df1.iloc[i,0]+b) >= df1.iloc[i,1]:
                df_TCC_new = pd.concat([df_TCC_new, df1.iloc[[i],:]], ignore_index=True)
        
        # the number of rows of df_TCC_new is the number of counts within the Gate
        cell_number_gate = len(df_TCC_new)
        
        # target volume for count in mL
        v = 1
        
        # factor for the conversion in cells per mL
        volume_factor = v/acq_vol
        
        # cell count per mL (TCC), includes conversion in mL and manual dilution factor
        TCC_per_mL = int(volume_factor * cell_number_gate) * dil_fac
        
        # calculate HNA/LNA cells/mL
        # HNA/LNA boundary on x-axis, value is minimum in histogram between peak regions for mineral water (Evian)
        x_HNA = x_LNA
        
        # create array for counting
        tcc_green_new = df_TCC_new.iloc[:,[0]].to_numpy()
        
        # create empty lists for HNA and LNA counts
        list_HNA = []
        list_LNA = []
        
        # fill the lists
        for l in range(len(tcc_green_new)):
            if tcc_green_new[l] > x_HNA:
                list_HNA.append(l)
            else:
                list_LNA.append(l)
                
        # cell count per mL (HNA and LNA), includes conversion in mL and manual dilution factor
        HNA_per_mL = int(volume_factor * len(list_HNA)) * dil_fac
        LNA_per_mL = int(volume_factor * len(list_LNA)) * dil_fac
        HNA_amount = HNA_per_mL / TCC_per_mL * 100
        
        print('\nTotal cells per mL:')
        print(TCC_per_mL)
        print('\nHNA cells per mL:')
        print(HNA_per_mL)
        print('\nLNA cells per mL:')
        print(LNA_per_mL)
        print('\nAmount of HNA cells:')
        print(HNA_amount)
              
        # =============================================================================
        #   Plot of the fcm fingerprint
        # =============================================================================
        # define font type
        plt.rcParams['font.family']='Times New Roman'
        
        # define the font style for tick notation
        def format_y_tick(x, pos):
            """Format y-axis tick labels in x10 notation."""
            if x == 0:
                return '0'
            exponent=int(np.log10(abs(x)))
            coefficient=x / 10 ** exponent
            if coefficient == 1:
                # added font family setting
                return r'$\mathregular{{10^{{\mathregular{{{}}}}}}}$'.format(exponent)
            elif coefficient.is_integer():
                # added font family setting
                return r'$\mathregular{{{:d}\times10^{{{}}}}}$'.format(int(coefficient), exponent)
            else:
                # added font family setting
                return r'$\mathregular{{{:0.1f}\times10^{{{}}}}}$'.format(coefficient, exponent)
        
        # chose colormap 
        magma_white_background = LinearSegmentedColormap.from_list('magma_white_background', [
            (0, '#ffffff'),    # White
            (1e-20, '#000003'), # Dark brown
            (0.2, '#530002'),   # Brown
            (0.4, '#bd2200'),   # Orange
            (0.6, '#fcaa00'),   # Yellow
            (0.8, '#fcf600'),   # Light yellow
            (1, '#ffffff')      # White
        ], N=256)
        
        # create the fcm fingerprint
        fig, ax = plt.subplots(1, 1, figsize=(8, 6), subplot_kw={'projection': 'scatter_density'})
        fig.subplots_adjust(left=0.15, right=0.95, bottom=0.15, top=0.95)
        density = ax.scatter_density(tcc_green, tcc_red, cmap=magma_white_background, vmin=0, vmax=10)
        # Colorbar with adjusted fontsize
        cbar = plt.colorbar(density, ax=ax)
        cbar.set_label('Number of points per pixel', fontsize=20, labelpad=10)  # Adjust the fontsize as needed
        cbar.ax.tick_params(axis='both', which='major', labelsize=20)
        # general settings
        ax.set_xlim(10,60000)
        ax.set_ylim(8,50000)
        ax = plt.yscale('log')
        ax = plt.xscale('log')
        ax = plt.xlabel('Green fluorescence in A.U.', fontsize=20, labelpad=10)
        ax = plt.ylabel('Red fluorescence in A.U.', fontsize=20, labelpad=10)
        ax = plt.title(name)
        ax = plt.xticks(fontsize=20)
        ax = plt.yticks(fontsize=20)
        # plot TCC gate
        ax = plt.plot(x_line,y_line, color='black')
        # plot LNA gate
        ax = plt.plot(x_line_LNA, y_line_LNA, color='black')    
        ax = plt.text(70, 11, 'LNA', fontsize=20, color='black')
        # plot HNA gate
        ax = plt.text(1000, 11, 'HNA', fontsize=20, color='black')
        plt.show()
            
        # =============================================================================
        #   Export results
        # =============================================================================
        # Create a pandas DataFrame containing the results
        df = pd.DataFrame({"Parameter": ["LNA", "HNA", "TCC", "HNA in %"],
                            "Counts [per mL]": [LNA_per_mL, HNA_per_mL, TCC_per_mL, HNA_amount],
                            "Counts [abs.]": [len(list_LNA), len(list_LNA), len(df_TCC_new), HNA_amount]})
        
        from datetime import datetime
        import openpyxl
        
        # Prompt the user to enter the file name and location
        file_name = filename
        file_path = location
        
        # Load the existing Excel workbook
        try:
            workbook = openpyxl.load_workbook(file_path + file_name)
        except FileNotFoundError:
            workbook = openpyxl.Workbook()
        
        # Set the active worksheet
        worksheet = workbook.active
        
        # If the worksheet doesn't already have headers, add them
        if worksheet.cell(row=1, column=1).value is None:
            worksheet.cell(row=1, column=1, value="Index")
            worksheet.cell(row=1, column=2, value="File Name")
            worksheet.cell(row=1, column=3, value="HNA")
            worksheet.cell(row=1, column=4, value="LNA")
            worksheet.cell(row=1, column=5, value="TCC")
            worksheet.cell(row=1, column=6, value="HNA in %")
            worksheet.cell(row=1, column=7, value="Time")
            worksheet.cell(row=1, column=8, value="Date")
            worksheet.cell(row=1, column=9, value='Sample ID')
        
        # Write the results to the Excel worksheet
        now = datetime.now()
        date_string = now.strftime("%Y-%m-%d")
        time_string = now.strftime("%H:%M:%S")
        filename_parts = name.split("__")
        index = filename_parts[-1].split(".")[0]
        
        number_rows_samples = worksheet.max_row
        
        #create a series with only sample id's to safe the rigth sample name in final excel
        def create_repeating_series(length, start_number=1):
        
            index_mapping = {1: 0, 2: 1, 3: 2}
            start_index = index_mapping[start_number]
        
            if sample_id == 3:
                repeating_sequence = cycle(range(1, 4))
                repeating_list = list(islice(repeating_sequence, start_index, start_index + length))
            elif sample_id == 2:
                repeating_sequence = cycle(range(1, 3))
                repeating_list = list(islice(repeating_sequence, start_index, start_index + length))
            elif sample_id == 1:
                repeating_list = sample_id_start
                #repeating_sequence = cycle(range(1, 2))
                #repeating_list = list(islice(repeating_sequence, start_index, start_index + length))
            return pd.Series(repeating_list, dtype=int)
        
        # Define the desired length of the series and the starting number
        length_parameter = number_rows_samples
        start_number_parameter = sample_id_start
        
        # Create the pandas Series
        result_series = create_repeating_series(length_parameter, start_number_parameter)
        
        # define the current sample bnumber
        sample_id_current = int(result_series.iloc[-1]) 
        
        # Write the results to the Excel worksheet
        worksheet.append([index, name, HNA_per_mL, LNA_per_mL, TCC_per_mL, HNA_amount, time_string, date_string, sample_id_current])
        
        # Adjust the width of the "File Name" column
        max_name_length = len(name)
        if max_name_length > 30:
            max_name_length = 30
        worksheet.column_dimensions['B'].width = max_name_length + 2
        
        # Adjust the width of the "File Name" column
        max_name_length_date_string = len(date_string)
        if max_name_length_date_string > 30:
            max_name_length_date_string = 30
        worksheet.column_dimensions['H'].width = max_name_length_date_string + 2
        
        # Save the Excel workbook
        workbook.save(file_path + file_name)
        
        # Print the table using tabulate
        print(tabulate(df, headers="keys", tablefmt="psql"))
        
        print("*********************************************************************************************************")
        print('Waiting for measurement')
        print("*********************************************************************************************************")
        
        # Wait for t seconds before checking for a new file
        time.sleep(t)