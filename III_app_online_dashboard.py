"""
@author: Jonas Schuster
"""
       
import plotly.graph_objects as go
import plotly.graph_objects as grob
#import jinja2
import dash    
from dash import dcc  
from dash import html       
from dash.dependencies import Input, Output
from dash.exceptions import PreventUpdate
import pandas as pd
import numpy as np
import openpyxl as xl;
from itertools import cycle, islice
import datetime


# =============================================================================
# Set input parameters for online analysis
# =============================================================================
sample_id = input("\n************************\nPlease select how many parallel samples are measured: 1, 2 or 3?\n************************\n")
sample_id = int(sample_id)
sample_id_start = input("\n************************\nSample ID of first Sample reading: 1, 2 or 3?\n************************\n")
sample_id_start = int(sample_id_start)
folder_fluorescence = input("\n************************\nEnter the location of the folder containing fluorescence data/results (model_data_6.xlsx) \n(e.g. C:/Users/Public/fluorescence_spectroscopy/:\n************************\n ")
folder_fcm = input("\n************************\nEnter the location of the folder containing flow cytometry data/results (fcm_results.xlsx) \n(e.g. C:/Users/Public/flow_cytometry/)\n************************\n ")
update = input("\n************************\nEnter the interval for dashboard updates in seconds\n************************\n ")

# Transformation from minutes to milliseconds
update_intervall = float(update) * 1000
print("************************************************")

# =============================================================================
# Set all general settings for dashboard
# =============================================================================
# chose style
external_stylesheets = ['https://codepen.io/chriddyp/pen/bWLwgP.css'] # other stylesheets to test: 'https://codepen.io/chriddyp/pen/brPBPO.css', 'https://codepen.io/chriddyp/pen/rzyyWo.css', 'https://codepen.io/chriddyp/pen/qaXKBR.css'

# initialize dash application
app = dash.Dash(__name__, external_stylesheets=external_stylesheets)

# define general colors
colors = {'background': '#111111', 'text': '#7FDBFF', 'marker': '#EE6AA7'}

#create a top-level 'html.Div' component that serves as the container for the other components with following properties
app.layout = html.Div([             

    html.Div(id="output-div",       
             style={'backgroundColor': colors['background']}, 
             children=[    
                 html.H1(
                     children='dashboard is loading, please wait...',
                     style={
                         'textAlign': 'center',
                         'color': colors['text']
                     }
                 )
                 ]),
   
    dcc.Interval(
                id='my_interval',
                disabled=False,     #if True, the counter will no longer update
                # define updating interval
                interval=update_intervall,    #increment the counter n_intervals every interval milliseconds
                n_intervals=0,      #number of times the interval has passed
                max_intervals=-1)    #number of times the interval will be fired.
                                    #if -1, then the interval has no limit (the default)
                                    #and if 0 then the interval stops running.
    ])

# =============================================================================
# generate fluorescence fingerprints (C1 - C6) based on 'base data' (72 samples)
# =============================================================================

# Read fingerprints of overall six component PARAFAC model, they don't change
fingerprints = pd.read_excel(folder_fluorescence+'fingerprints_model6.xlsx', sheet_name='Model6Loading')

# read ex and em wavelengths
ex = fingerprints['Ex'][[not pd.isnull(number) for number in fingerprints['Ex']]]
ex = ex.iloc[0:101].reset_index(drop=True)
em = fingerprints['Em'][[not pd.isnull(number) for number in fingerprints['Em']]]
em = em.iloc[0:100].reset_index(drop=True)

# read C1 score for all ex and em wavelengths
comp1_ex = fingerprints['Ex1'][[not pd.isnull(number) for number in fingerprints['Ex1']]]
comp1_ex = comp1_ex.iloc[0:101].reset_index(drop=True)
comp1_em = fingerprints['Em1'][[not pd.isnull(number) for number in fingerprints['Em1']]]
comp1_em = comp1_em.iloc[0:100].reset_index(drop=True)

# find ex index with max score (required afterwards)
# for later discrimation between C1 and C2 only focus on second peak, here: ignore first 10 rows
ind_max_ex_comp1 = np.argmax(comp1_ex.iloc[10:])+10

# find em index with max score (required afterwards)
ind_max_em_comp1 = np.argmax(comp1_em)

# read C2 score for all ex and em wavelengths
comp2_ex = fingerprints['Ex2'][[not pd.isnull(number) for number in fingerprints['Ex2']]]
comp2_ex = comp2_ex.iloc[0:101].reset_index(drop=True)
comp2_em = fingerprints['Em2'][[not pd.isnull(number) for number in fingerprints['Em2']]]
comp2_em = comp2_em.iloc[0:100].reset_index(drop=True)

# find ex and em index with max score (required afterwards)
ind_max_ex_comp2 = np.argmax(comp2_ex)
ind_max_em_comp2 = np.argmax(comp2_em)

# read C3 score for all ex and em wavelengths
comp3_ex = fingerprints['Ex3'][[not pd.isnull(number) for number in fingerprints['Ex3']]]
comp3_ex = comp3_ex.iloc[0:101].reset_index(drop=True)
comp3_em = fingerprints['Em3'][[not pd.isnull(number) for number in fingerprints['Em3']]]
comp3_em = comp3_em.iloc[0:100].reset_index(drop=True)

# find ex index with max score (required afterwards)
# for later discrimation between C3 and C4 only focus on second peak, here: ignore first 10 rows
ind_max_ex_comp3 = np.argmax(comp3_ex.iloc[10:])+10

# find em index with max score (required afterwards)
ind_max_em_comp3 = np.argmax(comp3_em)

# read C4 score for all ex and em wavelengths
comp4_ex = fingerprints['Ex4'][[not pd.isnull(number) for number in fingerprints['Ex4']]]
comp4_ex = comp4_ex.iloc[0:101].reset_index(drop=True)
comp4_em = fingerprints['Em4'][[not pd.isnull(number) for number in fingerprints['Em4']]]
comp4_em = comp4_em.iloc[0:100].reset_index(drop=True)

# find ex index with max score (required afterwards)
# for later discrimation between C4 and C3/C2 only focus on second peak, here: ignore first 10 rows
ind_max_ex_comp4 = np.argmax(comp4_ex.iloc[10:])+10

# find em index with max score (required afterwards)
ind_max_em_comp4 = np.argmax(comp4_em)

# read C5 score for all ex and em wavelengths
comp5_ex = fingerprints['Ex5'][[not pd.isnull(number) for number in fingerprints['Ex5']]]
comp5_ex = comp5_ex.iloc[0:101].reset_index(drop=True)
comp5_em = fingerprints['Em5'][[not pd.isnull(number) for number in fingerprints['Em5']]]
comp5_em = comp5_em.iloc[0:100].reset_index(drop=True)

# find ex and em index with max score (required afterwards)
ind_max_ex_comp5 = np.argmax(comp5_ex)
ind_max_em_comp5 = np.argmax(comp5_em)

# read C6 score for all ex and em wavelengths
comp6_ex = fingerprints['Ex6'][[not pd.isnull(number) for number in fingerprints['Ex6']]]
comp6_ex = comp6_ex.iloc[0:101].reset_index(drop=True)
comp6_em = fingerprints['Em6'][[not pd.isnull(number) for number in fingerprints['Em6']]]
comp6_em = comp6_em.iloc[0:100].reset_index(drop=True)

# find ex and em index with max score (required afterwards)
ind_max_ex_comp6 = np.argmax(comp6_ex)
ind_max_em_comp6 = np.argmax(comp6_em)

# create fingerprint array (Matrix based on ex and em intensities)
int_comp1 =  np.outer(comp1_ex, comp1_em)
int_comp2 =  np.outer(comp2_ex, comp2_em)
int_comp3 =  np.outer(comp3_ex, comp3_em)
int_comp4 =  np.outer(comp4_ex, comp4_em)
int_comp5 =  np.outer(comp5_ex, comp5_em)
int_comp6 =  np.outer(comp6_ex, comp6_em)

# =============================================================================
# plot fluorescence fingerprints
# =============================================================================

# C1
fp_c1 = grob.Figure(data=grob.Contour(x=em,y=ex,z=int_comp1, showscale=False, colorscale='Viridis'))
fp_c1.update_layout(
    title=dict(text='C1<br>allochthonous, terrestrial<br>humic-like, fulvic acid-like',xanchor='left'),
    xaxis=dict(title='emission wavelength in nm'),
    yaxis=dict(title='excitation wavelength in nm'),
    plot_bgcolor=colors['background'],
    paper_bgcolor=colors['background'],
    font_color=colors['text']
    )

# C2
fp_c2 = grob.Figure(data=grob.Contour(x=em,y=ex,z=int_comp2, showscale=False, colorscale='Viridis'))
fp_c2.update_layout(
    title=dict(text='C2<br>allochthonous and autochthonous, microbial and terrestrial<br>humic-like',xanchor='left'),
    xaxis=dict(title='emission wavelength in nm'),
    yaxis=dict(title='excitation wavelength in nm'),
    plot_bgcolor=colors['background'],
    paper_bgcolor=colors['background'],
    font_color=colors['text']
    )

# C3
fp_c3 = grob.Figure(data=grob.Contour(x=em,y=ex,z=int_comp3, showscale=False, colorscale='Viridis'))
fp_c3.update_layout(
    title=dict(text='C3<br>autochthonous, microbial<br>humic-like',xanchor='left'),
    xaxis=dict(title='emission wavelength in nm'),
    yaxis=dict(title='excitation wavelength in nm'),
    plot_bgcolor=colors['background'],
    paper_bgcolor=colors['background'],
    font_color=colors['text']
    )

# C4
fp_c4 = grob.Figure(data=grob.Contour(x=em,y=ex,z=int_comp4, showscale=False, colorscale='Viridis'))
fp_c4.update_layout(
    title=dict(text='C4<br>microbial and terrestrial<br>humic-like, fulvic acid-like',xanchor='left'),
    xaxis=dict(title='emission wavelength in nm'),
    yaxis=dict(title='excitation wavelength in nm'),
    plot_bgcolor=colors['background'],
    paper_bgcolor=colors['background'],
    font_color=colors['text']
    )

# C5
fp_c5 = grob.Figure(data=grob.Contour(x=em,y=ex,z=int_comp5, showscale=False, colorscale='Viridis'))
fp_c5.update_layout(
    title=dict(text='C5<br>terrestrial<br>humic-like, humic acid-like',xanchor='left'),
    xaxis=dict(title='emission wavelength in nm'),
    yaxis=dict(title='excitation wavelength in nm'),
    plot_bgcolor=colors['background'],
    paper_bgcolor=colors['background'],
    font_color=colors['text']
    )

# C6
fp_c6 = grob.Figure(data=grob.Contour(x=em,y=ex,z=int_comp6, showscale=False, colorscale='Viridis'))
fp_c6.update_layout(
    title=dict(text='C6<br>microbially produced, autochthonous<br>amino acid-like, protein-like, tryptophan-like',xanchor='left'),
    xaxis=dict(title='emission wavelength in nm'),
    yaxis=dict(title='excitation wavelength in nm'),
    plot_bgcolor=colors['background'],
    paper_bgcolor=colors['background'],
    font_color=colors['text']
    )

# Generate figures to check order of components
print('\nPlease check if the declarations of the component numbers are right.\n')
fp_c1.show(renderer="png")
fp_c2.show(renderer="png")
fp_c3.show(renderer="png")
fp_c4.show(renderer="png")
fp_c5.show(renderer="png")
fp_c6.show(renderer="png")

# Prompt user to confirm
confirmation = input("Are all declarations right? (Press Enter to continue) \n")

# =============================================================================
# call and update dashboard
# =============================================================================
@app.callback(Output(component_id="output-div", component_property="children"),
              Input(component_id="my_interval", component_property="n_intervals")
              )

def update(num):              # num = n_intervals
    """update every x (interval) seconds"""
    if num==0:
        raise PreventUpdate
    else:
        import datetime
        # Following starts when there is a new reading
# =============================================================================
#       Import latest fluorescence measurement and identify components and its ex and em max
# =============================================================================
        # read new fluorescence results, excel is updatet every time Aqualog has a new reading
        loading = pd.read_excel(folder_fluorescence+'model_data_6.xlsx', sheet_name='Model6Loading')

        # read new C1 (Fmax1) score for all ex and em wavelengths 
        compFmax1_ex = loading['Ex1'][[not pd.isnull(number) for number in loading['Ex1']]]
        compFmax1_ex = compFmax1_ex.iloc[0:101].reset_index(drop=True)
        compFmax1_em = loading['Em1'][[not pd.isnull(number) for number in loading['Em1']]]
        compFmax1_em = compFmax1_em.iloc[0:100].reset_index(drop=True)
        
        # find first ex max
        max1 = compFmax1_ex.idxmax()
        
        # create a new series excluding first ex max to find second ex max, remove 13 neighbours from the first ex max
        neighbours_to_remove = 13
        
        # Define the range of indices (ex) to exclude
        start_index = max(max1 - neighbours_to_remove, 0)
        end_index = min(max1 + neighbours_to_remove +1, len(compFmax1_ex))
        
        # create new series without first max for identifying second ex max
        compFmax1_ex_2 = pd.concat([compFmax1_ex[:start_index], compFmax1_ex[end_index:]])

        # identify index for first ex max
        ind_max_ex_cFmax1_1 = compFmax1_ex.idxmax()
        
        # identify index for second ex max
        ind_max_ex_cFmax1_2 = compFmax1_ex_2.idxmax()

        # identify em index of maximal value
        ind_max_em_cFmax1 = np.argmax(compFmax1_em)
       
        # read new C2 (Fmax2) score for all ex and em wavelengths 
        compFmax2_ex = loading['Ex2'][[not pd.isnull(number) for number in loading['Ex2']]]
        compFmax2_ex = compFmax2_ex.iloc[0:101].reset_index(drop=True)
        compFmax2_em = loading['Em2'][[not pd.isnull(number) for number in loading['Em2']]]
        compFmax2_em = compFmax2_em.iloc[0:100].reset_index(drop=True)
        
        # find first ex max
        max1 = compFmax2_ex.idxmax()
        
        # create a new series excluding first ex max to find second ex max, remove 13 neighbours from the first ex max
        neighbours_to_remove = 13
        
        # Define the range of indices (ex) to exclude
        start_index = max(max1 - neighbours_to_remove, 0)
        end_index = min(max1 + neighbours_to_remove +1, len(compFmax2_ex))
        
        # create new series without first max for identifying second ex max     
        compFmax2_ex_2 = pd.concat([compFmax2_ex[:start_index], compFmax2_ex[end_index:]])
        
        # identify index for first ex max
        ind_max_ex_cFmax2_1 = compFmax2_ex.idxmax()
        
        # identify index for second ex max
        ind_max_ex_cFmax2_2 = compFmax2_ex_2.idxmax()
        
        # identify em index of maximal value
        ind_max_em_cFmax2 = np.argmax(compFmax2_em)
        
        # read new C3 (Fmax3) score for all ex and em wavelengths
        compFmax3_ex = loading['Ex3'][[not pd.isnull(number) for number in loading['Ex3']]]
        compFmax3_ex = compFmax3_ex.iloc[0:101].reset_index(drop=True)
        compFmax3_em = loading['Em3'][[not pd.isnull(number) for number in loading['Em3']]]
        compFmax3_em = compFmax3_em.iloc[0:100].reset_index(drop=True)

        # find first ex max        
        max1 = compFmax3_ex.idxmax()
        
        # create a new series excluding first ex max to find second ex max, remove 13 neighbours from the first ex max
        neighbours_to_remove = 13
        
        # Define the range of indices (ex) to exclude
        start_index = max(max1 - neighbours_to_remove, 0)
        end_index = min(max1 + neighbours_to_remove +1, len(compFmax3_ex))
        
        # create new series without first max for identifying second ex max
        compFmax3_ex_2 = pd.concat([compFmax3_ex[:start_index], compFmax3_ex[end_index:]])
        
        # identify index for first ex max
        ind_max_ex_cFmax3_1 = compFmax3_ex.idxmax()
        
        # identify index for second ex max
        ind_max_ex_cFmax3_2 = compFmax3_ex_2.idxmax()
        
        # identify em index of maximal value
        ind_max_em_cFmax3 = np.argmax(compFmax3_em)
        
        # read new C4 (Fmax4) score for all ex and em wavelengths
        compFmax4_ex = loading['Ex4'][[not pd.isnull(number) for number in loading['Ex4']]]
        compFmax4_ex = compFmax4_ex.iloc[0:101].reset_index(drop=True)
        compFmax4_em = loading['Em4'][[not pd.isnull(number) for number in loading['Em4']]]
        compFmax4_em = compFmax4_em.iloc[0:100].reset_index(drop=True)
        
        # find first ex max
        max1 = compFmax4_ex.idxmax()
        
        # create a new series excluding first ex max to find second ex max, remove 13 neighbours from the first ex max
        neighbours_to_remove = 13
        
        # Define the range of indices (ex) to exclude
        start_index = max(max1 - neighbours_to_remove, 0)
        end_index = min(max1 + neighbours_to_remove +1, len(compFmax4_ex))
        
        # create new series without first max for identifying second ex max
        compFmax4_ex_2 = pd.concat([compFmax4_ex[:start_index], compFmax4_ex[end_index:]])
        
        # identify index for first ex max
        ind_max_ex_cFmax4_1 = compFmax4_ex.idxmax()
        
        # identify index for second ex max
        ind_max_ex_cFmax4_2 = compFmax4_ex_2.idxmax()
        
        # identify em index of maximal value
        ind_max_em_cFmax4 = np.argmax(compFmax4_em)
        
        # read new C5 (Fmax5) score for all ex and em wavelengths
        compFmax5_ex = loading['Ex5'][[not pd.isnull(number) for number in loading['Ex5']]]
        compFmax5_ex = compFmax5_ex.iloc[0:101].reset_index(drop=True)
        compFmax5_em = loading['Em5'][[not pd.isnull(number) for number in loading['Em5']]]
        compFmax5_em = compFmax5_em.iloc[0:100].reset_index(drop=True)
        
        # find first ex max
        max1 = compFmax5_ex.idxmax()
        
        # create a new series excluding first ex max to find second ex max, remove 13 neighbours from the first ex max
        neighbours_to_remove = 13
        
        # Define the range of indices (ex) to exclude
        start_index = max(max1 - neighbours_to_remove, 0)
        end_index = min(max1 + neighbours_to_remove +1, len(compFmax5_ex))
        
        # create new series without first max for identifying second ex max
        compFmax5_ex_2 = pd.concat([compFmax5_ex[:start_index], compFmax5_ex[end_index:]])
        
        # identify index for first ex max
        ind_max_ex_cFmax5_1 = compFmax5_ex.idxmax()
        
        # identify index for second ex max
        ind_max_ex_cFmax5_2 = compFmax5_ex_2.idxmax()
        
        # identify em index of maximal value
        ind_max_em_cFmax5 = np.argmax(compFmax5_em)
        
        # read new C6 (Fmax6) score for all ex and em wavelengths
        compFmax6_ex = loading['Ex6'][[not pd.isnull(number) for number in loading['Ex6']]]
        compFmax6_ex = compFmax6_ex.iloc[0:101].reset_index(drop=True)
        compFmax6_em = loading['Em6'][[not pd.isnull(number) for number in loading['Em6']]]
        compFmax6_em = compFmax6_em.iloc[0:100].reset_index(drop=True)
        
        # find first ex max
        max1 = compFmax6_ex.idxmax()
        
        # create a new series excluding first ex max to find second ex max, remove 13 neighbours from the first ex max
        neighbours_to_remove = 13
        
        # Define the range of indices (ex) to exclude
        start_index = max(max1 - neighbours_to_remove, 0)
        end_index = min(max1 + neighbours_to_remove +1, len(compFmax6_ex))
        
        # create new series without first max for identifying second ex max
        compFmax6_ex_2 = pd.concat([compFmax6_ex[:start_index], compFmax6_ex[end_index:]])
        
        # identify index for first ex max
        ind_max_ex_cFmax6_1 = compFmax6_ex.idxmax()
        
        # identify index for second ex max
        ind_max_ex_cFmax6_2 = compFmax6_ex_2.idxmax()
        
        # identify em index of maximal value
        ind_max_em_cFmax6 = np.argmax(compFmax6_em)


# =============================================================================
#       Recap of ex and em max from 'base data'
# =============================================================================
        # C1, Lambda_(0/67), ex:250/370 em:466 humic-like, fulvic acid like     -> column 1
        # C2, Lambda_(1/63), ex:250/325 em:450 humic-like                       -> column 2
        # C3, Lambda_(18/48), ex:250/305 em:400 humic-like                      -> column 3
        # C4, Lambda_(33/54), ex:250/350 em:420 humic-like, fulvic acid like    -> column 4
        # C5, Lambda_(0/86), ex:250/275/420 em:525 Humic-like, humic acid like  -> column 5
        # C6, Lambda_(9/30), ex: 280 em: 335 amino acid-like                    -> column 6

# =============================================================================
#         Preparation step
#         
#         Compare latest reading ('model_data_6') components with 'base data' components (ex/em max)
#         # Index range for ex and em max of fingerprints, based on experience
#         C1: ex: 35-41, em: 61-69
#         C2: ex: 0-6, em:61-68
#         C3: ex: 13-20, em:46-52 
#         C4: ex:30-36, em:49-55  
#         C5: ex:0-6, em:84-90
#         C6: ex:6-13, em:27-34
# =============================================================================
    
        # maxima from 'base data' fingerprint model (here: fingerprints_model6.xlsx)
        d = {'Component':['C1', 'C2', 'C3', 'C4', 'C5', 'C6'], 'Ex':[ind_max_ex_comp1, ind_max_ex_comp2, ind_max_ex_comp3, ind_max_ex_comp4,
                                                                     ind_max_ex_comp5, ind_max_ex_comp6], 'Em':[ind_max_em_comp1, ind_max_em_comp2,
                                                                    ind_max_em_comp3, ind_max_em_comp4, ind_max_em_comp5, ind_max_em_comp6]}
        df_ind_max_fingerprint = pd.DataFrame(data=d)
        
        # maxima from last reading sample (here: model_data_6.xlsx)
        ind_max = {'Name':['Fmax1_1', 'Fmax1_2','Fmax2_1','Fmax2_2','Fmax3_1','Fmax3_2','Fmax4_1','Fmax4_2','Fmax5_1','Fmax5_2','Fmax6_1','Fmax6_2'],
                   'Ex':[ind_max_ex_cFmax1_1, ind_max_ex_cFmax1_2, ind_max_ex_cFmax2_1,ind_max_ex_cFmax2_2 ,ind_max_ex_cFmax3_1,ind_max_ex_cFmax3_2,ind_max_ex_cFmax4_1,ind_max_ex_cFmax4_2, ind_max_ex_cFmax5_1,ind_max_ex_cFmax5_2, ind_max_ex_cFmax6_1, ind_max_ex_cFmax6_2],
                   'Em':[ind_max_em_cFmax1, ind_max_em_cFmax1, ind_max_em_cFmax2, ind_max_em_cFmax2, ind_max_em_cFmax3, ind_max_em_cFmax3, ind_max_em_cFmax4, ind_max_em_cFmax4, ind_max_em_cFmax5, ind_max_em_cFmax5, ind_max_em_cFmax6, ind_max_em_cFmax6]}
        df_ind_max_current = pd.DataFrame(data=ind_max)
        
        # adapt number of 'base data' samples (here: 72)
        base_samples = len(np.array(range(1,73))) 
        # create dataframe with intensities from last reading (dataframe: loading)
        entfnan = loading[['i','Fmax1','Fmax2','Fmax3','Fmax4','Fmax5','Fmax6']].dropna()
        base_df = entfnan.tail(base_samples)
        base_df.reset_index(inplace=True, drop=True)
        base_df = base_df.drop(base_df.columns[0], axis=1)
        new_column = pd.Series(range(1,base_samples+1))
        base_df['i'] = new_column

        #create target dataframe
        d_sorted = ['i', 'C1', 'C2', 'C3', 'C4', 'C5', 'C6']
        base_df_sorted = pd.DataFrame(columns = d_sorted)
        base_df_sorted['i'] = np.array(range(1,73))
        

        # check which ex/em max standard is equal to the latest reading one and returns the component
        # cFmax1 = component x 
        df_C1_ex = []
        df_C1_em = []
    
        # check which component from the latest reading belongs to C1 
        # the wavelength ranges are predefined and latest reading is checked therefore
        for j in range(35, 41):     # between Ex-Max = 355 - 375 nm
            # those of the latest reading that have an ex max here are returned
            C1_Ex = df_ind_max_current.loc[df_ind_max_current['Ex'] == j,'Name']
            df_C1_ex.append(C1_Ex)
            res_C1_ex = pd.concat(df_C1_ex) 
            
        # same for em max
        for i in range(61, 69):
            C1_Em = df_ind_max_current.loc[df_ind_max_current['Em'] == i,'Name']
            df_C1_em.append(C1_Em)
            res_C1_em = pd.concat(df_C1_em) 
            
        # check which Fmax (latest reading) is ex and em max
        C1 = list(set(res_C1_ex).intersection(res_C1_em))
        #C1 = str(C1).replace('[','').replace(']','').replace('\'','').replace('\"','')   
        #C1 = C1[:-2]
        
        # same procedure for the remaining components
        df_C2_ex = []
        df_C2_em = []

        for j in range(0,6):     
            C2_Ex = df_ind_max_current.loc[df_ind_max_current['Ex'] == j,'Name']
            df_C2_ex.append(C2_Ex)
            res_C2_ex = pd.concat(df_C2_ex)  

        for i in range(60,68):
            C2_Em = df_ind_max_current.loc[df_ind_max_current['Em'] == i,'Name']
            df_C2_em.append(C2_Em)
            res_C2_em = pd.concat(df_C2_em) 

        C2 = list(set(res_C2_ex).intersection(res_C2_em))
        #C2 = str(C2).replace('[','').replace(']','').replace('\'','').replace('\"','') 
        #C2 = C2[:-2]

        df_C3_ex = []
        df_C3_em = []

        for j in range(13,20):     
            C3_Ex = df_ind_max_current.loc[df_ind_max_current['Ex'] == j,'Name']
            df_C3_ex.append(C3_Ex)
            res_C3_ex = pd.concat(df_C3_ex)  

        for i in range(43,53):
            C3_Em = df_ind_max_current.loc[df_ind_max_current['Em'] == i,'Name']
            df_C3_em.append(C3_Em)
            res_C3_em = pd.concat(df_C3_em) 

        C3 = list(set(res_C3_ex).intersection(res_C3_em))
        #C3 = str(C3).replace('[','').replace(']','').replace('\'','').replace('\"','')  
        #C3 = C3[:-2]

        df_C4_ex = []
        df_C4_em = []

        for j in range(30,36):     
            C4_Ex = df_ind_max_current.loc[df_ind_max_current['Ex'] == j,'Name']
            df_C4_ex.append(C4_Ex)
            res_C4_ex = pd.concat(df_C4_ex)  

        for i in range(49,55):
            C4_Em = df_ind_max_current.loc[df_ind_max_current['Em'] == i,'Name']
            df_C4_em.append(C4_Em)
            res_C4_em = pd.concat(df_C4_em) 

        C4 = list(set(res_C4_ex).intersection(res_C4_em))
        #C4 = str(C4).replace('[','').replace(']','').replace('\'','').replace('\"','')   
        #C4 = C4[:-2]

        df_C5_ex = []
        df_C5_em = []

        for j in range(0,6):     
            C5_Ex = df_ind_max_current.loc[df_ind_max_current['Ex'] == j,'Name']
            df_C5_ex.append(C5_Ex)
            res_C5_ex = pd.concat(df_C5_ex)  

        for i in range(84,90):
            C5_Em = df_ind_max_current.loc[df_ind_max_current['Em'] == i,'Name']
            df_C5_em.append(C5_Em)
            res_C5_em = pd.concat(df_C5_em) 

        C5 = list(set(res_C5_ex).intersection(res_C5_em))
        #C5 = str(C5).replace('[','').replace(']','').replace('\'','').replace('\"','')   
        #C5 = C5[:-2]

        df_C6_ex = []
        df_C6_em = []

        for j in range(6,13):     
            C6_Ex = df_ind_max_current.loc[df_ind_max_current['Ex'] == j,'Name']
            df_C6_ex.append(C6_Ex)
            res_C6_ex = pd.concat(df_C6_ex)  

        for i in range(28,34):
            C6_Em = df_ind_max_current.loc[df_ind_max_current['Em'] == i,'Name']
            df_C6_em.append(C6_Em)
            res_C6_em = pd.concat(df_C6_em) 

        C6 = list(set(res_C6_ex).intersection(res_C6_em))
        #C6 = str(C6).replace('[','').replace(']','').replace('\'','').replace('\"','') 
        #C6 = C6[:-2]
       
        # create list with all Components
        Comp_list = [C1,C2,C3,C4,C5,C6]
        Comp_list_call = []
        
        # loop through each component list to remove second index 
        for component in Comp_list:
            modified_sublist = [item[:-2] if len(item) > 1 else item for item in component]
            Comp_list_call.append(modified_sublist)
        
        Comp_list_flat = [item for sublist in Comp_list_call for item in sublist]
        
        # identify duplicate index
        Comp_duplicates = []
        Comp_seen = set()

        for item in Comp_list_flat:
            if item in Comp_seen:
                if item not in Comp_duplicates:
                    Comp_duplicates.append(item)
            else:
                Comp_seen.add(item)
               
        # remove duplicate index
        Comp_value_to_remove = Comp_duplicates
        Comp_modified_list = []
        Comp_modified_list_new = []
        
        # remove all duplicate index
        Comp_modified_list = [[item for item in sublist if item not in Comp_value_to_remove] for sublist in Comp_list_call]
        
        # add the index where there is no value
        Comp_modified_list_new = [sublist if sublist else Comp_value_to_remove for sublist in Comp_modified_list]
        
        C1 = str(Comp_modified_list_new[0]).replace('[','').replace(']','').replace('\'','').replace('\"','')
        C2 = str(Comp_modified_list_new[1]).replace('[','').replace(']','').replace('\'','').replace('\"','')
        C3 = str(Comp_modified_list_new[2]).replace('[','').replace(']','').replace('\'','').replace('\"','')
        C4 = str(Comp_modified_list_new[3]).replace('[','').replace(']','').replace('\'','').replace('\"','')
        C5 = str(Comp_modified_list_new[4]).replace('[','').replace(']','').replace('\'','').replace('\"','')
        C6 = str(Comp_modified_list_new[5]).replace('[','').replace(']','').replace('\'','').replace('\"','')
        
        # check in which column FmaxX is in base_df 
        columnIndex1 = base_df.columns.get_loc(C1)
        # Safe this column into new variable  
        l1 = base_df.iloc[0:,columnIndex1] 
        # And copy this to the C1 column of the sorted dataframe
        base_df_sorted['C1'] = pd.Series(l1)
        columnIndex2 = base_df.columns.get_loc(C2)
        l2 = base_df.iloc[0:,columnIndex2] 
        base_df_sorted['C2'] = pd.Series(l2)
        columnIndex3 = base_df.columns.get_loc(C3)
        l3 = base_df.iloc[0:,columnIndex3] 
        base_df_sorted['C3'] = pd.Series(l3)
        columnIndex4 = base_df.columns.get_loc(C4)
        l4 = base_df.iloc[0:,columnIndex4] 
        base_df_sorted['C4'] = pd.Series(l4)
        columnIndex5 = base_df.columns.get_loc(C5)
        l5 = base_df.iloc[0:,columnIndex5] 
        base_df_sorted['C5'] = pd.Series(l5)
        columnIndex6 = base_df.columns.get_loc(C6)
        l6 = base_df.iloc[0:,columnIndex6] 
        base_df_sorted['C6'] = pd.Series(l6)

# =============================================================================
#         # plot base data scores
#         base_c1 = px.scatter(base_df_sorted, x="i", y="C1")
#         base_c2 = px.scatter(base_df_sorted, x="i", y="C2")
#         base_c3 = px.scatter(base_df_sorted, x="i", y="C3")
#         base_c4 = px.scatter(base_df_sorted, x="i", y="C4")
#         base_c5 = px.scatter(base_df_sorted, x="i", y="C5")
#         base_c6 = px.scatter(base_df_sorted, x="i", y="C6")
#         
#         #inline plots
#         base_c1.show(renderer='png')
#         base_c2.show(renderer='png')
#         base_c3.show(renderer='png')
#         base_c4.show(renderer='png')
#         base_c5.show(renderer='png')
#         base_c6.show(renderer='png')
# =============================================================================
        
# =============================================================================
#       # prepare added dataset (all measurements excluding 'base data')
# =============================================================================
        # read score excel
        scores = pd.read_excel(folder_fluorescence+'scores.xlsx', sheet_name='Model6Loading')
        # identify number of measurements 
        added_samples = loading["i"].notna().sum() - base_samples

        # copy latest sample from 'model_data_6' to destination excel-file ('scores')Neueste hinzugefuegte Probe wird vom source zum destination Excel-File kopiert
        # latest reading Excel-File
        folder_excel_s= folder_fluorescence+'model_data_6.xlsx'
        wb_s = xl.load_workbook(folder_excel_s)
        ws_s = wb_s.worksheets[1] #loading sheet 2: "Model6Loading"
        ws_s_date = wb_s.worksheets[0] #loading ist sheet 1: "Model6Report"
        
        # Destination Excel-File
        folder_excel_d = folder_fluorescence+'scores.xlsx'
        wb_d = xl.load_workbook(folder_excel_d)
        ws_d = wb_d.worksheets[0] 

        # determine the number of rows of 'score' 
        max_row_d = ws_d.max_row  
            
        # copy last row of 'scores'
        # columnIndex is column of 'model_data_6', Fmax index is already in the right position
        # copy the datetime  of the last reading
        date_new = ws_s_date.cell(5, 3).value
        
        # convert string to a datetime object
        datetime_obj = datetime.datetime.strptime(date_new, "%d-%b-%Y %H:%M:%S")
        
        # Convert the datetime object to the desired format
        date_new = datetime_obj.strftime('%Y-%m-%d %H:%M:%S')        
        
        # scores of already existing readings
        old_C1 = round(scores.iloc[max_row_d-2, 1],10)
        old_C2 = round(scores.iloc[max_row_d-2, 2],10)
        old_C3 = round(scores.iloc[max_row_d-2, 3],10)
        old_C4 = round(scores.iloc[max_row_d-2, 4],10)
        old_C5 = round(scores.iloc[max_row_d-2, 5],10)
        old_C6 = round(scores.iloc[max_row_d-2, 6],10)
        
        # check if there is an initial reading
        if added_samples == 0:
            print('Waiting for first Aqualog (Fluorescence) Sample')
            added_C1 = old_C1
            added_C2 = old_C2
            added_C3 = old_C3
            added_C4 = old_C4
            added_C5 = old_C5
            added_C6 = old_C6
        else:
            # variable columnIndex contains information about the right column, regarding the former excecuted calculations of peak distinguishment
            added_C1 = round(ws_s.cell(added_samples+1, columnIndex1+2).value,10)
            added_C2 = round(ws_s.cell(added_samples+1, columnIndex2+2).value,10)
            added_C3 = round(ws_s.cell(added_samples+1, columnIndex3+2).value,10)
            added_C4 = round(ws_s.cell(added_samples+1, columnIndex4+2).value,10)
            added_C5 = round(ws_s.cell(added_samples+1, columnIndex5+2).value,10)
            added_C6 = round(ws_s.cell(added_samples+1, columnIndex6+2).value,10)
        
        # check if the latest reading in 'model_data_6' is the same as the latest reaading in 'scores'
        # app.py is continuously updatet
        # first row is header
        if added_C1 == old_C1 and added_C2 == old_C2 and added_C3 == old_C3 and added_C4 == old_C4 and added_C5 == old_C5 and added_C6 == old_C6:
            print("Waiting for new Aqualog (Fluorescence) Sample")
        else:
            # latest reading in 'model_data_6' varies from the latest reading in 'scores'
            print('New Aqualog (Fluorescence) Sample')
            
            # latest reading in 'model_data_6' is added to a new row in 'scores' 
            ws_d.cell(max_row_d+1, 2).value = added_C1
            ws_d.cell(max_row_d+1, 3).value = added_C2
            ws_d.cell(max_row_d+1, 4).value = added_C3
            ws_d.cell(max_row_d+1, 5).value = added_C4
            ws_d.cell(max_row_d+1, 6).value = added_C5
            ws_d.cell(max_row_d+1, 7).value = added_C6 
            
            # add datetime
            ws_d.cell(max_row_d+1, 8).value = date_new
            
            # add current sample ID
            length_parameter = max_row_d-base_samples
            start_number_parameter = sample_id_start
            def create_repeating_series(length, start_number=1):
                 index_mapping = {1: 0, 2: 1, 3: 2}
                 start_index = index_mapping[start_number]
                 if sample_id == 3:
                     repeating_sequence = cycle(range(1, 4))
                     repeating_list = list(islice(repeating_sequence, start_index, start_index + length))
                 elif sample_id == 2:
                     repeating_sequence = cycle(range(1, 3))
                     repeating_list = list(islice(repeating_sequence, start_index, start_index + length))
                 elif sample_id == 1:
                     repeating_list = sample_id_start
                 return pd.Series(repeating_list, dtype=int)
            
            # Create the pandas Series with sample ID
            result_series = create_repeating_series(length_parameter, start_number_parameter)
            sample_id_current = int(result_series.iloc[-1]) 
            ws_d.cell(max_row_d+1, 9).value = sample_id_current
            
            # for-loop to include numeration in first column 
            for j in range(1, max_row_d+1):
                ws_d.cell(row = j+1, column = 1).value = j 
            
            # save new 'scores' as an excel
            wb_d.save(str(folder_excel_d))  
                           
        # read new 'scores' 
        added_df = pd.read_excel(folder_fluorescence+'scores.xlsx')
        
        # create df for different sample ID (1 - 3)
        added_df_1 = added_df[added_df['Sample ID'] == 1]
        added_df_2 = added_df[added_df['Sample ID'] == 2]
        added_df_3 = added_df[added_df['Sample ID'] == 3]
        
        # calculate mean and STD. Threshold factor from literature 3 (10.1016/j.watres.2021.117387)
        t = 3
        
        # Take first 20 readings to create baseline, this can be also felxible (10, 20, 30, 40...)
        
        # Sample 1 mean values
        C1_mean_s1 = added_df_1.iloc[0:19, 1:2].mean()
        C2_mean_s1 = added_df_1.iloc[0:19, 2:3].mean()
        C3_mean_s1 = added_df_1.iloc[0:19, 3:4].mean()
        C4_mean_s1 = added_df_1.iloc[0:19, 4:5].mean()
        C5_mean_s1 = added_df_1.iloc[0:19, 5:6].mean()
        C6_mean_s1 = added_df_1.iloc[0:19, 6:7].mean()
        
        # Sample 1 Std values
        C1_std_s1 = added_df_1.iloc[0:19, 1:2].std()
        C2_std_s1 = added_df_1.iloc[0:19, 2:3].std()
        C3_std_s1 = added_df_1.iloc[0:19, 3:4].std()
        C4_std_s1 = added_df_1.iloc[0:19, 4:5].std()
        C5_std_s1 = added_df_1.iloc[0:19, 5:6].std()
        C6_std_s1 = added_df_1.iloc[0:19, 6:7].std()
        
        # sample 1 Define upper and lower limits for respective baselines
        C1_s1_base_upper_limit = (C1_mean_s1 + t * C1_std_s1)
        C1_s1_base_lower_limit = (C1_mean_s1 - t * C1_std_s1)
        C2_s1_base_upper_limit = (C2_mean_s1 + t * C2_std_s1)
        C2_s1_base_lower_limit = (C2_mean_s1 - t * C2_std_s1)
        C3_s1_base_upper_limit = (C3_mean_s1 + t * C3_std_s1)
        C3_s1_base_lower_limit = (C3_mean_s1 - t * C3_std_s1)
        C4_s1_base_upper_limit = (C4_mean_s1 + t * C4_std_s1)
        C4_s1_base_lower_limit = (C4_mean_s1 - t * C4_std_s1)
        C5_s1_base_upper_limit = (C5_mean_s1 + t * C5_std_s1)
        C5_s1_base_lower_limit = (C5_mean_s1 - t * C5_std_s1)
        C6_s1_base_upper_limit = (C6_mean_s1 + t * C6_std_s1)
        C6_s1_base_lower_limit = (C6_mean_s1 - t * C6_std_s1)
        
        # Sample 2 mean values
        C1_mean_s2 = added_df_2.iloc[0:19, 1:2].mean()
        C2_mean_s2 = added_df_2.iloc[0:19, 2:3].mean()
        C3_mean_s2 = added_df_2.iloc[0:19, 3:4].mean()
        C4_mean_s2 = added_df_2.iloc[0:19, 4:5].mean()
        C5_mean_s2 = added_df_2.iloc[0:19, 5:6].mean()
        C6_mean_s2 = added_df_2.iloc[0:19, 6:7].mean()
        
        # Sample 2 Std values
        C1_std_s2 = added_df_2.iloc[0:19, 1:2].std()
        C2_std_s2 = added_df_2.iloc[0:19, 2:3].std()
        C3_std_s2 = added_df_2.iloc[0:19, 3:4].std()
        C4_std_s2 = added_df_2.iloc[0:19, 4:5].std()
        C5_std_s2 = added_df_2.iloc[0:19, 5:6].std()
        C6_std_s2 = added_df_2.iloc[0:19, 6:7].std()
        
        # sample 2 Define upper and lower limits for respective baselines
        C1_s2_base_upper_limit = (C1_mean_s2 + t * C1_std_s2)
        C1_s2_base_lower_limit = (C1_mean_s2 - t * C1_std_s2)
        C2_s2_base_upper_limit = (C2_mean_s2 + t * C2_std_s2)
        C2_s2_base_lower_limit = (C2_mean_s2 - t * C2_std_s2)
        C3_s2_base_upper_limit = (C3_mean_s2 + t * C3_std_s2)
        C3_s2_base_lower_limit = (C3_mean_s2 - t * C3_std_s2)
        C4_s2_base_upper_limit = (C4_mean_s2 + t * C4_std_s2)
        C4_s2_base_lower_limit = (C4_mean_s2 - t * C4_std_s2)
        C5_s2_base_upper_limit = (C5_mean_s2 + t * C5_std_s2)
        C5_s2_base_lower_limit = (C5_mean_s2 - t * C5_std_s2)
        C6_s2_base_upper_limit = (C6_mean_s2 + t * C6_std_s2)
        C6_s2_base_lower_limit = (C6_mean_s2 - t * C6_std_s2)
        
        # Sample 3 mean values
        C1_mean_s3 = added_df_3.iloc[0:19, 1:2].mean()
        C2_mean_s3 = added_df_3.iloc[0:19, 2:3].mean()
        C3_mean_s3 = added_df_3.iloc[0:19, 3:4].mean()
        C4_mean_s3 = added_df_3.iloc[0:19, 4:5].mean()
        C5_mean_s3 = added_df_3.iloc[0:19, 5:6].mean()
        C6_mean_s3 = added_df_3.iloc[0:19, 6:7].mean()
        
        # Sample 3 Std values
        C1_std_s3 = added_df_3.iloc[0:19, 1:2].std()
        C2_std_s3 = added_df_3.iloc[0:19, 2:3].std()
        C3_std_s3 = added_df_3.iloc[0:19, 3:4].std()
        C4_std_s3 = added_df_3.iloc[0:19, 4:5].std()
        C5_std_s3 = added_df_3.iloc[0:19, 5:6].std()
        C6_std_s3 = added_df_3.iloc[0:19, 6:7].std()
        
        # sample 3 Define upper and lower limits for respective baselines
        C1_s3_base_upper_limit = (C1_mean_s3 + t * C1_std_s3)
        C1_s3_base_lower_limit = (C1_mean_s3 - t * C1_std_s3)
        C2_s3_base_upper_limit = (C2_mean_s3 + t * C2_std_s3)
        C2_s3_base_lower_limit = (C2_mean_s3 - t * C2_std_s3)
        C3_s3_base_upper_limit = (C3_mean_s3 + t * C3_std_s3)
        C3_s3_base_lower_limit = (C3_mean_s3 - t * C3_std_s3)
        C4_s3_base_upper_limit = (C4_mean_s3 + t * C4_std_s3)
        C4_s3_base_lower_limit = (C4_mean_s3 - t * C4_std_s3)
        C5_s3_base_upper_limit = (C5_mean_s3 + t * C5_std_s3)
        C5_s3_base_lower_limit = (C5_mean_s3 - t * C5_std_s3)
        C6_s3_base_upper_limit = (C6_mean_s3 + t * C6_std_s3)
        C6_s3_base_lower_limit = (C6_mean_s3 - t * C6_std_s3)
    
# =============================================================================
#       # plot fluorescence data (dashboard update)
# =============================================================================
        
# =============================================================================
#       # C1
# =============================================================================
        fig_c1 = go.Figure()
        
        fig_c1.add_trace(
            go.Scatter(x=added_df_1['Date/Time'], y=added_df_1['C1'], name='sample 1', line=dict(color=colors['marker']))
            )
        fig_c1.add_trace(
            go.Scatter(x=added_df_2['Date/Time'], y=added_df_2['C1'], name='sample 2', line=dict(color='cyan'))
            )
        fig_c1.add_trace(
            go.Scatter(x=added_df_3['Date/Time'], y=added_df_3['C1'], name='sample 3', line=dict(color='yellow'))
            )
        
        # Sample 1 lower Limit
        fig_c1.add_trace(
            go.Scatter(x=list(added_df_1['Date/Time']), y = [C1_s1_base_lower_limit[0]] * len(added_df_1.index), 
                       name='upper / lower \nlimit sample 1', mode='lines',
                       visible=False, line=dict(color=colors['marker'], dash='dash'))
            )
        
        # Sample 1 Upper Limit
        fig_c1.add_trace(
            go.Scatter(x=list(added_df_1['Date/Time']), y = [C1_s1_base_upper_limit[0]] * len(added_df_1.index), 
                       mode='lines', showlegend=False,
                       visible=False, line=dict(color=colors['marker'], dash='dash'))
            )
        
        # Sample 2 lower Limit
        fig_c1.add_trace(
            go.Scatter(x=list(added_df_2['Date/Time']), y = [C1_s2_base_lower_limit[0]] * len(added_df_2.index), 
                       name='upper / lower \nlimit sample 2', mode='lines',
                       visible=False, line=dict(color='cyan', dash='dash'))
            )
        
        # Sample 2 Upper Limit
        fig_c1.add_trace(
            go.Scatter(x=list(added_df_2['Date/Time']), y = [C1_s2_base_upper_limit[0]] * len(added_df_2.index), 
                       mode='lines', showlegend=False,
                       visible=False, line=dict(color='cyan', dash='dash'))
            )
        
        # Sample 3 lower Limit
        fig_c1.add_trace(
            go.Scatter(x=list(added_df_3['Date/Time']), y = [C1_s3_base_lower_limit[0]] * len(added_df_3.index), 
                       name='upper / lower \nlimit sample 3', mode='lines',
                       visible=False, line=dict(color='yellow', dash='dash'))
            )
        
        # Sample 3 Upper Limit
        fig_c1.add_trace(
            go.Scatter(x=list(added_df_3['Date/Time']), y = [C1_s3_base_upper_limit[0]] * len(added_df_3.index), 
                       mode='lines', showlegend=False,
                       visible=False, line=dict(color='yellow', dash='dash'))
            )
        
        #fig_c1.show(renderer='png') 
        
        # C1 general settings 
        fig_c1.update_layout(
            updatemenus=[
                dict(
                    active=0,
                    buttons=list([
                        dict(label='all samples',
                             method='update',
                             args=[{'visible': [True, True, True, False, False, False, False, False, False]},
                                   {'title': 'C1 fluorescence scores of all samples'}]),
                        dict(label='sample 1',
                             method='update',
                             args=[{'visible': [True, False, False, True, True, False, False, False, False]},
                                   {'title': 'C1 fluorescence scores of sample 1'}]),
                        dict(label='sample 2',
                             method='update',
                             args=[{'visible': [False, True, False, False, False, True, True, False, False]},
                                   {'title': 'C1 fluorescence scores of sample 2'}]),
                        dict(label='sample 3',
                             method='update',
                             args=[{'visible': [False, False, True, False, False, False, False, True, True]},
                                   {'title': 'C1 fluorescence scores of sample 3'}])
                        
                        ])
                    )
                ],
            showlegend=True
            )
       
        fig_c1.update_layout(
            plot_bgcolor=colors['background'],
            xaxis=dict(title='date/time', 
                       showgrid=False, 
                       linewidth=2, 
                       linecolor=colors['text'],
                       zeroline=False
                       ),
            yaxis=dict(title='fluorescence score in r.u.',
                       # dtick=0.02,
                       # range=[0.1, 0.30], 
                       # 
                       showgrid=False, 
                       linewidth=2, 
                       linecolor=colors['text'],
                       zeroline=False
                       ),
            paper_bgcolor=colors['background'],
            font_color=colors['text']
        )
        
        #fig_c1.show(renderer='png')
        
# =============================================================================
#       # C2
# =============================================================================
        fig_c2 = go.Figure()
        
        fig_c2.add_trace(
            go.Scatter(x=added_df_1['Date/Time'], y=added_df_1['C2'], name='sample 1', line=dict(color=colors['marker']))
            )
        fig_c2.add_trace(
            go.Scatter(x=added_df_2['Date/Time'], y=added_df_2['C2'], name='sample 2', line=dict(color='cyan'))
            )
        fig_c2.add_trace(
            go.Scatter(x=added_df_3['Date/Time'], y=added_df_3['C2'], name='sample 3', line=dict(color='yellow'))
            )
        
        # Sample 1 lower Limit
        fig_c2.add_trace(
            go.Scatter(x=list(added_df_1['Date/Time']), y = [C2_s1_base_lower_limit[0]] * len(added_df_1.index), 
                       name='upper / lower \nlimit sample 1', mode='lines',
                       visible=False, line=dict(color=colors['marker'], dash='dash'))
            )
        
        # Sample 1 Upper Limit
        fig_c2.add_trace(
            go.Scatter(x=list(added_df_1['Date/Time']), y = [C2_s1_base_upper_limit[0]] * len(added_df_1.index), 
                       mode='lines', showlegend=False,
                       visible=False, line=dict(color=colors['marker'], dash='dash'))
            )
        
        # Sample 2 lower Limit
        fig_c2.add_trace(
            go.Scatter(x=list(added_df_2['Date/Time']), y = [C2_s2_base_lower_limit[0]] * len(added_df_2.index), 
                       name='upper / lower \nlimit sample 2', mode='lines',
                       visible=False, line=dict(color='cyan', dash='dash'))
            )
        
        # Sample 2 Upper Limit
        fig_c2.add_trace(
            go.Scatter(x=list(added_df_2['Date/Time']), y = [C2_s2_base_upper_limit[0]] * len(added_df_2.index), 
                       mode='lines', showlegend=False,
                       visible=False, line=dict(color='cyan', dash='dash'))
            )
        
        # Sample 3 lower Limit
        fig_c2.add_trace(
            go.Scatter(x=list(added_df_3['Date/Time']), y = [C2_s3_base_lower_limit[0]] * len(added_df_3.index), 
                       name='upper / lower \nlimit sample 3', mode='lines',
                       visible=False, line=dict(color='yellow', dash='dash'))
            )
        
        # Sample 3 Upper Limit
        fig_c2.add_trace(
            go.Scatter(x=list(added_df_3['Date/Time']), y = [C2_s3_base_upper_limit[0]] * len(added_df_3.index), 
                       mode='lines', showlegend=False,
                       visible=False, line=dict(color='yellow', dash='dash'))
            )
        
        #fig_c2.show(renderer='png') 
        
        # C2 general settings
        fig_c2.update_layout(
            updatemenus=[
                dict(
                    active=0,
                    buttons=list([
                        dict(label='all Samples',
                             method='update',
                             args=[{'visible': [True, True, True, False, False, False, False, False, False]},
                                   {'title': 'C2 fluorescence scores of all samples'}]),
                        dict(label='Sample 1',
                             method='update',
                             args=[{'visible': [True, False, False, True, True, False, False, False, False]},
                                   {'title': 'C2 fluorescence scores of sample 1'}]),
                        dict(label='Sample 2',
                             method='update',
                             args=[{'visible': [False, True, False, False, False, True, True, False, False]},
                                   {'title': 'C2 fluorescence scores of sample 2'}]),
                        dict(label='Sample 3',
                             method='update',
                             args=[{'visible': [False, False, True, False, False, False, False, True, True]},
                                   {'title': 'C2 fluorescence scores of sample 3'}])
                        
                        ])
                    )
                ],
            showlegend=True
            )
        
        fig_c2.update_layout(
            plot_bgcolor=colors['background'],
            xaxis=dict(title='date/time', 
                       showgrid=False, 
                       linewidth=2, 
                       linecolor=colors['text'],
                       zeroline=False
                       ),
            yaxis=dict(title='fluorescence score in r.u.',
                       # dtick=0.05,
                       # range=[0.1, 0.4], 
                       showgrid=False, 
                       linewidth=2, 
                       linecolor=colors['text'],
                       zeroline=False
                       ),
            paper_bgcolor=colors['background'],
            font_color=colors['text']
        )
        
        #fig_c2.show(renderer='png')
        
# =============================================================================
#         # C3
# =============================================================================
        fig_c3 = go.Figure()
        
        fig_c3.add_trace(
            go.Scatter(x=added_df_1['Date/Time'], y=added_df_1['C3'], name='sample 1', line=dict(color=colors['marker']))
            )
        fig_c3.add_trace(
            go.Scatter(x=added_df_2['Date/Time'], y=added_df_2['C3'], name='sample 2', line=dict(color='cyan'))
            )
        fig_c3.add_trace(
            go.Scatter(x=added_df_3['Date/Time'], y=added_df_3['C3'], name='sample 3', line=dict(color='yellow'))
            )
        
        # Sample 1 lower Limit
        fig_c3.add_trace(
            go.Scatter(x=list(added_df_1['Date/Time']), y = [C3_s1_base_lower_limit[0]] * len(added_df_1.index), 
                       name='upper / lower \nlimit sample 1', mode='lines',
                       visible=False, line=dict(color=colors['marker'], dash='dash'))
            )
        
        # Sample 1 Upper Limit
        fig_c3.add_trace(
            go.Scatter(x=list(added_df_1['Date/Time']), y = [C3_s1_base_upper_limit[0]] * len(added_df_1.index), 
                       mode='lines', showlegend=False,
                       visible=False, line=dict(color=colors['marker'], dash='dash'))
            )
        
        # Sample 2 lower Limit
        fig_c3.add_trace(
            go.Scatter(x=list(added_df_2['Date/Time']), y = [C3_s2_base_lower_limit[0]] * len(added_df_2.index), 
                       name='upper / lower \nlimit sample 2', mode='lines',
                       visible=False, line=dict(color='cyan', dash='dash'))
            )
        
        # Sample 2 Upper Limit
        fig_c3.add_trace(
            go.Scatter(x=list(added_df_2['Date/Time']), y = [C3_s2_base_upper_limit[0]] * len(added_df_2.index), 
                       mode='lines', showlegend=False,
                       visible=False, line=dict(color='cyan', dash='dash'))
            )
        
        # Sample 3 lower Limit
        fig_c3.add_trace(
            go.Scatter(x=list(added_df_3['Date/Time']), y = [C3_s3_base_lower_limit[0]] * len(added_df_3.index), 
                       name='upper / lower \nlimit sample 3', mode='lines',
                       visible=False, line=dict(color='yellow', dash='dash'))
            )
        
        # Sample 3 Upper Limit
        fig_c3.add_trace(
            go.Scatter(x=list(added_df_3['Date/Time']), y = [C3_s3_base_upper_limit[0]] * len(added_df_3.index), 
                       mode='lines', showlegend=False,
                       visible=False, line=dict(color='yellow', dash='dash'))
            )
        
        #fig_c3.show(renderer='png') 
        
        # C3 general setting
        fig_c3.update_layout(
            updatemenus=[
                dict(
                    active=0,
                    buttons=list([
                        dict(label='all samples',
                             method='update',
                             args=[{'visible': [True, True, True, False, False, False, False, False, False]},
                                   {'title': 'C3 fluorescence scores of all samples'}]),
                        dict(label='Sample 1',
                             method='update',
                             args=[{'visible': [True, False, False, True, True, False, False, False, False]},
                                   {'title': 'C3 fluorescence scores of sample 1'}]),
                        dict(label='Sample 2',
                             method='update',
                             args=[{'visible': [False, True, False, False, False, True, True, False, False]},
                                   {'title': 'C3 fluorescence scores of sample 2'}]),
                        dict(label='Sample 3',
                             method='update',
                             args=[{'visible': [False, False, True, False, False, False, False, True, True]},
                                   {'title': 'C3 sluorescence scores of sample 3'}])
                        
                        ])
                    )
                ],
            showlegend=True
            )
        
        fig_c3.update_layout(
            plot_bgcolor=colors['background'],
            xaxis=dict(title='date/time', 
                       showgrid=False, 
                       linewidth=2, 
                       linecolor=colors['text'],
                       zeroline=False
                       ),
            yaxis=dict(title='fluorescence score in r.u.',
                       # dtick=0.05,
                       # range=[0, 0.2], 
                       showgrid=False, 
                       linewidth=2, 
                       linecolor=colors['text'],
                       zeroline=False
                       ),
            paper_bgcolor=colors['background'],
            font_color=colors['text']
        )
        
        #fig_c3.show(renderer='png')
        
# =============================================================================
#         # C4
# =============================================================================
        fig_c4 = go.Figure()
        
        fig_c4.add_trace(
            go.Scatter(x=added_df_1['Date/Time'], y=added_df_1['C4'], name='sample 1', line=dict(color=colors['marker']))
            )
        fig_c4.add_trace(
            go.Scatter(x=added_df_2['Date/Time'], y=added_df_2['C4'], name='sample 2', line=dict(color='cyan'))
            )
        fig_c4.add_trace(
            go.Scatter(x=added_df_3['Date/Time'], y=added_df_3['C4'], name='sample 3', line=dict(color='yellow'))
            )
        
        # Sample 1 lower Limit
        fig_c4.add_trace(
            go.Scatter(x=list(added_df_1['Date/Time']), y = [C4_s1_base_lower_limit[0]] * len(added_df_1.index), 
                       name='upper / lower \nlimit sample 1', mode='lines',
                       visible=False, line=dict(color=colors['marker'], dash='dash'))
            )
        
        # Sample 1 Upper Limit
        fig_c4.add_trace(
            go.Scatter(x=list(added_df_1['Date/Time']), y = [C4_s1_base_upper_limit[0]] * len(added_df_1.index), 
                       mode='lines', showlegend=False,
                       visible=False, line=dict(color=colors['marker'], dash='dash'))
            )
        
        # Sample 2 lower Limit
        fig_c4.add_trace(
            go.Scatter(x=list(added_df_2['Date/Time']), y = [C4_s2_base_lower_limit[0]] * len(added_df_2.index), 
                       name='upper / lower \nlimit sample 2', mode='lines',
                       visible=False, line=dict(color='cyan', dash='dash'))
            )
        
        # Sample 2 Upper Limit
        fig_c4.add_trace(
            go.Scatter(x=list(added_df_2['Date/Time']), y = [C4_s2_base_upper_limit[0]] * len(added_df_2.index), 
                       mode='lines', showlegend=False,
                       visible=False, line=dict(color='cyan', dash='dash'))
            )
        
        # Sample 3 lower Limit
        fig_c4.add_trace(
            go.Scatter(x=list(added_df_3['Date/Time']), y = [C4_s3_base_lower_limit[0]] * len(added_df_3.index), 
                       name='upper / lower \nlimit sample 3', mode='lines',
                       visible=False, line=dict(color='yellow', dash='dash'))
            )
        
        # Sample 3 Upper Limit
        fig_c4.add_trace(
            go.Scatter(x=list(added_df_3['Date/Time']), y = [C4_s3_base_upper_limit[0]] * len(added_df_3.index), 
                       mode='lines', showlegend=False,
                       visible=False, line=dict(color='yellow', dash='dash'))
            )
        
        #fig_c4.show(renderer='png')
        
        # C4 general settings      
        fig_c4.update_layout(
            updatemenus=[
                dict(
                    active=0,
                    buttons=list([
                        dict(label='all Samples',
                             method='update',
                             args=[{'visible': [True, True, True, False, False, False, False, False, False]},
                                   {'title': 'C4 fluorescence scores of all samples'}]),
                        dict(label='Sample 1',
                             method='update',
                             args=[{'visible': [True, False, False, True, True, False, False, False, False]},
                                   {'title': 'C4 fluorescence scores of sample 1'}]),
                        dict(label='Sample 2',
                             method='update',
                             args=[{'visible': [False, True, False, False, False, True, True, False, False]},
                                   {'title': 'C4 fluorescence scores of sample 2'}]),
                        dict(label='Sample 3',
                             method='update',
                             args=[{'visible': [False, False, True, False, False, False, False, True, True]},
                                   {'title': 'C4 fluorescence scores of sample 3'}])
                        
                        ])
                    )
                ],
            showlegend=True
            )
        
        fig_c4.update_layout(
            plot_bgcolor=colors['background'],
            xaxis=dict(title='date/time', 
                       showgrid=False, 
                       linewidth=2, 
                       linecolor=colors['text'],
                       zeroline=False
                       ),
            yaxis=dict(title='fluorescence score in r.u.',
                       # dtick=0.05,
                       # range=[0.0, 0.3], 
                       showgrid=False, 
                       linewidth=2, 
                       linecolor=colors['text'],
                       zeroline=False
                       ),
            paper_bgcolor=colors['background'],
            font_color=colors['text']
        )
        
        #fig_c4.show(renderer='png')
        
# =============================================================================
#         # C5
# =============================================================================
        fig_c5 = go.Figure()
        
        fig_c5.add_trace(
            go.Scatter(x=added_df_1['Date/Time'], y=added_df_1['C5'], name='sample 1', line=dict(color=colors['marker']))
            )
        fig_c5.add_trace(
            go.Scatter(x=added_df_2['Date/Time'], y=added_df_2['C5'], name='sample 2', line=dict(color='cyan'))
            )
        fig_c5.add_trace(
            go.Scatter(x=added_df_3['Date/Time'], y=added_df_3['C5'], name='sample 3', line=dict(color='yellow'))
            )
        
        # Sample 1 lower Limit
        fig_c5.add_trace(
            go.Scatter(x=list(added_df_1['Date/Time']), y = [C5_s1_base_lower_limit[0]] * len(added_df_1.index), 
                       name='upper / lower \nlimit sample 1', mode='lines',
                       visible=False, line=dict(color=colors['marker'], dash='dash'))
            )
        
        # Sample 1 Upper Limit
        fig_c5.add_trace(
            go.Scatter(x=list(added_df_1['Date/Time']), y = [C5_s1_base_upper_limit[0]] * len(added_df_1.index), 
                       mode='lines', showlegend=False,
                       visible=False, line=dict(color=colors['marker'], dash='dash'))
            )
        
        # Sample 2 lower Limit
        fig_c5.add_trace(
            go.Scatter(x=list(added_df_2['Date/Time']), y = [C5_s2_base_lower_limit[0]] * len(added_df_2.index), 
                       name='upper / lower \nlimit sample 2', mode='lines',
                       visible=False, line=dict(color='cyan', dash='dash'))
            )
        
        # Sample 2 Upper Limit
        fig_c5.add_trace(
            go.Scatter(x=list(added_df_2['Date/Time']), y = [C5_s2_base_upper_limit[0]] * len(added_df_2.index), 
                       mode='lines', showlegend=False,
                       visible=False, line=dict(color='cyan', dash='dash'))
            )
        
        # Sample 3 lower Limit
        fig_c5.add_trace(
            go.Scatter(x=list(added_df_3['Date/Time']), y = [C5_s3_base_lower_limit[0]] * len(added_df_3.index), 
                       name='upper / lower \nlimit sample 3', mode='lines',
                       visible=False, line=dict(color='yellow', dash='dash'))
            )
        
        # Sample 3 Upper Limit
        fig_c5.add_trace(
            go.Scatter(x=list(added_df_3['Date/Time']), y = [C5_s3_base_upper_limit[0]] * len(added_df_3.index), 
                       mode='lines', showlegend=False,
                       visible=False, line=dict(color='yellow', dash='dash'))
            )
        
        #fig_c5.show(renderer='png') 
        
        # C5 general settings
        fig_c5.update_layout(
            updatemenus=[
                dict(
                    active=0,
                    buttons=list([
                        dict(label='all samples',
                             method='update',
                             args=[{'visible': [True, True, True, False, False, False, False, False, False]},
                                   {'title': 'C5 fluorescence scores of all samples'}]),
                        dict(label='Sample 1',
                             method='update',
                             args=[{'visible': [True, False, False, True, True, False, False, False, False]},
                                   {'title': 'C5 fluorescence scores of sample 1'}]),
                        dict(label='Sample 2',
                             method='update',
                             args=[{'visible': [False, True, False, False, False, True, True, False, False]},
                                   {'title': 'C5 fluorescence scores of sample 2'}]),
                        dict(label='Sample 3',
                             method='update',
                             args=[{'visible': [False, False, True, False, False, False, False, True, True]},
                                   {'title': 'C5 fluorescence scores of sample 3'}])
                        
                        ])
                    )
                ],
            showlegend=True
            )
        
        fig_c5.update_layout(
            plot_bgcolor=colors['background'],
            xaxis=dict(title='date/time', 
                       showgrid=False, 
                       linewidth=2, 
                       linecolor=colors['text'],
                       zeroline=False
                       ),
            yaxis=dict(title='fluorescence score in r.u.',
                       # dtick=0.05,
                       # range=[0.0, 0.3], 
                       showgrid=False, 
                       linewidth=2, 
                       linecolor=colors['text'],
                       zeroline=False
                       ),
            paper_bgcolor=colors['background'],
            font_color=colors['text']
        )
        
        #fig_c5.show(renderer='png')
        
# =============================================================================
#         # C6
# =============================================================================
        fig_c6 = go.Figure()
        
        fig_c6.add_trace(
            go.Scatter(x=added_df_1['Date/Time'], y=added_df_1['C6'], name='sample 1', line=dict(color=colors['marker']))
            )
        fig_c6.add_trace(
            go.Scatter(x=added_df_2['Date/Time'], y=added_df_2['C6'], name='sample 2', line=dict(color='cyan'))
            )
        fig_c6.add_trace(
            go.Scatter(x=added_df_3['Date/Time'], y=added_df_3['C6'], name='sample 3', line=dict(color='yellow'))
            )
        
        # Sample 1 lower Limit
        fig_c6.add_trace(
            go.Scatter(x=list(added_df_1['Date/Time']), y = [C6_s1_base_lower_limit[0]] * len(added_df_1.index), 
                       name='upper / lower \nlimit sample 1', mode='lines',
                       visible=False, line=dict(color=colors['marker'], dash='dash'))
            )
        
        # Sample 1 Upper Limit
        fig_c6.add_trace(
            go.Scatter(x=list(added_df_1['Date/Time']), y = [C6_s1_base_upper_limit[0]] * len(added_df_1.index), 
                       mode='lines', showlegend=False,
                       visible=False, line=dict(color=colors['marker'], dash='dash'))
            )
        
        # Sample 2 lower Limit
        fig_c6.add_trace(
            go.Scatter(x=list(added_df_2['Date/Time']), y = [C6_s2_base_lower_limit[0]] * len(added_df_2.index), 
                       name='upper / lower \nlimit sample 2', mode='lines',
                       visible=False, line=dict(color='cyan', dash='dash'))
            )
        
        # Sample 2 Upper Limit
        fig_c6.add_trace(
            go.Scatter(x=list(added_df_2['Date/Time']), y = [C6_s2_base_upper_limit[0]] * len(added_df_2.index), 
                       mode='lines', showlegend=False,
                       visible=False, line=dict(color='cyan', dash='dash'))
            )
        
        # Sample 3 lower Limit
        fig_c6.add_trace(
            go.Scatter(x=list(added_df_3['Date/Time']), y = [C6_s3_base_lower_limit[0]] * len(added_df_3.index), 
                       name='upper / lower \nlimit sample 3', mode='lines',
                       visible=False, line=dict(color='yellow', dash='dash'))
            )
        
        # Sample 3 Upper Limit
        fig_c6.add_trace(
            go.Scatter(x=list(added_df_3['Date/Time']), y = [C6_s3_base_upper_limit[0]] * len(added_df_3.index), 
                       mode='lines', showlegend=False,
                       visible=False, line=dict(color='yellow', dash='dash'))
            )
        
        #fig_c6.show(renderer='png') 
        
        # C6 general settings
        fig_c6.update_layout(
            updatemenus=[
                dict(
                    active=0,
                    buttons=list([
                        dict(label='all samples',
                             method='update',
                             args=[{'visible': [True, True, True, False, False, False, False, False, False]},
                                   {'title': 'C6 fluorescence scores of all samples'}]),
                        dict(label='Sample 1',
                             method='update',
                             args=[{'visible': [True, False, False, True, True, False, False, False, False]},
                                   {'title': 'C6 fluorescence scores of sample 1'}]),
                        dict(label='Sample 2',
                             method='update',
                             args=[{'visible': [False, True, False, False, False, True, True, False, False]},
                                   {'title': 'C6 fluorescence scores of sample 2'}]),
                        dict(label='Sample 3',
                             method='update',
                             args=[{'visible': [False, False, True, False, False, False, False, True, True]},
                                   {'title': 'C6 fluorescence scores of sample 3'}])
                        
                        ])
                    )
                ],
            showlegend=True
            )
        
        fig_c6.update_layout(
            plot_bgcolor=colors['background'],
            xaxis=dict(title='date/time', 
                       showgrid=False, 
                       linewidth=2, 
                       linecolor=colors['text'],
                       zeroline=False
                       ),
            yaxis=dict(title='fluorescence score in r.u.',
                       # dtick=0.05,
                       # range=[0, 0.3], 
                       showgrid=False, 
                       linewidth=2, 
                       linecolor=colors['text'],
                       zeroline=False
                       ),
            paper_bgcolor=colors['background'],
            font_color=colors['text']
        )
        
        #fig_c6.show(renderer='png') 
        
# =============================================================================
#         # read and safe SAC data
# =============================================================================          
        # read sac excel ('scores_sak254' and 'scores_sak436')
        sak254_old = pd.read_excel(folder_fluorescence+'scores_sak254.xlsx')
        sak436_old = pd.read_excel(folder_fluorescence+'scores_sak436.xlsx')
        
        # transform to float values
        sak254_old_float = float(sak254_old.iloc[-1,0])
        sak436_old_float = float(sak436_old.iloc[-1])
        
        # read latest sac reading ('sak254' and 'sak436')
        sak254_new = pd.read_csv(folder_fluorescence+'sak254.txt', header=None, names=['SAK 254'])      
        sak436_new = pd.read_csv(folder_fluorescence+'sak436.txt', header=None, names=['SAK 436'])
        
        # transform to float value / Multiply with 100 because of the unit conversion 1/cm to 1/m
        sak254_new_float = float(sak254_new.iloc[-1,])*100
        sak436_new_float = float(sak436_new.iloc[-1,])*100
        
        # check if there is a new reading
        if round(sak254_old_float, 10) == round(sak254_new_float, 10) and round(sak436_old_float, 10) == round(sak436_new_float, 10):
            print('Waiting for new Aqualog (Absorbance) Sample')
            sak = pd.concat([sak254_old, sak436_old], axis = 1)
                 
        else:
            print('New Aqualog (Absorbance) Sample')
            
            # safe new sac readings in excels
            sak254 = pd.concat([sak254_old, sak254_new*100], ignore_index=True)
            
            # read time stamp from added_df and add it to sac excels
            datetime = added_df.iloc[-1:,7:8]
            sak254.iloc[-1:,1:2] = datetime
            sak254.iloc[-1:,3:4] = sample_id_current
            sak254.to_excel(folder_fluorescence+'scores_sak254.xlsx', index=None)
            sak436 = pd.concat([sak436_old, sak436_new*100], ignore_index=True)
            sak436.to_excel(folder_fluorescence+'scores_sak436.xlsx', index=None)
            sak = pd.concat([sak254, sak436], axis=1)
              
        sak_added = sak.iloc[1:,:]
        
# =============================================================================
#         # plot sac data (dashboard update)
# =============================================================================
        # create df for different Sample ID (1 - 3)
        sak_1 = sak_added[sak_added['Sample ID'] == 1] 
        sak_2 = sak_added[sak_added['Sample ID'] == 2]
        sak_3 = sak_added[sak_added['Sample ID'] == 3]
        
        # Sample 1 mean values
        sak254_mean_s1 = sak_1.iloc[0:19, 0:1].mean()
        sak436_mean_s1 = sak_1.iloc[0:19, 4:5].mean()
        
        # Sample 1 Std values
        sak254_std_s1 = sak_1.iloc[0:19, 0:1].std()
        sak436_std_s1 = sak_1.iloc[0:19, 4:5].std()
        
        # sample 1 Define upper and lower limits for respective baselines
        sak254_s1_base_upper_limit = (sak254_mean_s1 + t * sak254_std_s1)
        sak254_s1_base_lower_limit = (sak254_mean_s1 - t * sak254_std_s1)
        sak436_s1_base_upper_limit = (sak436_mean_s1 + t * sak436_std_s1)
        sak436_s1_base_lower_limit = (sak436_mean_s1 - t * sak436_std_s1)
        
        # Sample 2 mean values
        sak254_mean_s2 = sak_2.iloc[0:19, 0:1].mean()
        sak436_mean_s2 = sak_2.iloc[0:19, 4:5].mean()
        
        # Sample 2 Std values
        sak254_std_s2 = sak_2.iloc[0:19, 0:1].std()
        sak436_std_s2 = sak_2.iloc[0:19, 4:5].std()
        
        # sample 2 Define upper and lower limits for respective baselines
        sak254_s2_base_upper_limit = (sak254_mean_s2 + t * sak254_std_s2)
        sak254_s2_base_lower_limit = (sak254_mean_s2 - t * sak254_std_s2)
        sak436_s2_base_upper_limit = (sak436_mean_s2 + t * sak436_std_s2)
        sak436_s2_base_lower_limit = (sak436_mean_s2 - t * sak436_std_s2)
        
        # Sample 3 mean values
        sak254_mean_s3 = sak_3.iloc[0:19, 0:1].mean()
        sak436_mean_s3 = sak_3.iloc[0:19, 4:5].mean()
        
        # Sample 3 Std values
        sak254_std_s3 = sak_3.iloc[0:19, 0:1].std()
        sak436_std_s3 = sak_3.iloc[0:19, 4:5].std()
        
        # sample 3 Define upper and lower limits for respective baselines
        sak254_s3_base_upper_limit = (sak254_mean_s3 + t * sak254_std_s3)
        sak254_s3_base_lower_limit = (sak254_mean_s3 - t * sak254_std_s3)
        sak436_s3_base_upper_limit = (sak436_mean_s3 + t * sak436_std_s3)
        sak436_s3_base_lower_limit = (sak436_mean_s3 - t * sak436_std_s3)
        
# =============================================================================
#         # sac 254
# =============================================================================
        fig_sak254 = go.Figure()
        
        fig_sak254.add_trace(
            go.Scatter(x=sak_1['Date/Time'], y=sak_1['SAK 254'], name='Sample 1', line=dict(color=colors['marker']))
            )
        fig_sak254.add_trace(
            go.Scatter(x=sak_2['Date/Time'], y=sak_2['SAK 254'], name='Sample 2', line=dict(color='cyan'))
            )
        fig_sak254.add_trace(
            go.Scatter(x=sak_3['Date/Time'], y=sak_3['SAK 254'], name='Sample 3', line=dict(color='yellow'))
            )
        
        # Sample 1 lower Limit
        fig_sak254.add_trace(
            go.Scatter(x=list(sak_1['Date/Time']), y = [sak254_s1_base_lower_limit[0]] * len(sak_1.index), 
                       name='Upper / Lower \nLimit Sample 1', mode='lines',
                       visible=False, line=dict(color=colors['marker'], dash='dash'))
            )
        
        # Sample 1 Upper Limit
        fig_sak254.add_trace(
            go.Scatter(x=list(sak_1['Date/Time']), y = [sak254_s1_base_upper_limit[0]] * len(sak_1.index), 
                       mode='lines', showlegend=False,
                       visible=False, line=dict(color=colors['marker'], dash='dash'))
            )
        
        # Sample 2 lower Limit
        fig_sak254.add_trace(
            go.Scatter(x=list(sak_2['Date/Time']), y = [sak254_s2_base_lower_limit[0]] * len(sak_2.index), 
                       name='Upper / Lower \nLimit Sample 2', mode='lines',
                       visible=False, line=dict(color='cyan', dash='dash'))
            )
        
        # Sample 2 Upper Limit
        fig_sak254.add_trace(
            go.Scatter(x=list(sak_2['Date/Time']), y = [sak254_s2_base_upper_limit[0]] * len(sak_2.index), 
                       mode='lines', showlegend=False,
                       visible=False, line=dict(color='cyan', dash='dash'))
            )
        
        # Sample 3 lower Limit
        fig_sak254.add_trace(
            go.Scatter(x=list(sak_3['Date/Time']), y = [sak254_s3_base_lower_limit[0]] * len(sak_3.index), 
                       name='Upper / Lower \nLimit Sample 3', mode='lines',
                       visible=False, line=dict(color='yellow', dash='dash'))
            )
        
        # Sample 3 Upper Limit
        fig_sak254.add_trace(
            go.Scatter(x=list(sak_3['Date/Time']), y = [sak254_s3_base_upper_limit[0]] * len(sak_3.index), 
                       mode='lines', showlegend=False,
                       visible=False, line=dict(color='yellow', dash='dash'))
            )
        
        #fig_sak254.show(renderer='png') 
        
        # sac254 general settings
        fig_sak254.update_layout(
            updatemenus=[
                dict(
                    active=0,
                    buttons=list([
                        dict(label='All Samples',
                             method='update',
                             args=[{'visible': [True, True, True, False, False, False, False, False, False]},
                                   {'title': 'All Samples'}]),
                        dict(label='Sample 1',
                             method='update',
                             args=[{'visible': [True, False, False, True, True, False, False, False, False]},
                                   {'title': 'Sample 1'}]),
                        dict(label='Sample 2',
                             method='update',
                             args=[{'visible': [False, True, False, False, False, True, True, False, False]},
                                   {'title': 'Sample 2'}]),
                        dict(label='Sample 3',
                             method='update',
                             args=[{'visible': [False, False, True, False, False, False, False, True, True]},
                                   {'title': 'Sample 3'}])
                        
                        ])
                    )
                ],
            showlegend=True
            )
        
        fig_sak254.update_layout(
            plot_bgcolor=colors['background'],
            xaxis=dict(title='date/time', 
                       showgrid=False, 
                       linewidth=2, 
                       #linecolor=colors['text'],
                       zeroline=False
                       ),
            yaxis=dict(title='spectral absorbance in 1/m',
                       # dtick=0.02,
                       # range=[0.1, 0.30], 
                       # 
                       showgrid=False, 
                       linewidth=2, 
                       #linecolor=colors['text'],
                       zeroline=False
                       ),
            paper_bgcolor=colors['background'],
            font_color=colors['text']
        )
        
        #fig_sak254.show(renderer='png')
        
# =============================================================================
#         # sac 436 
# ============================================================================
        fig_sak436 = go.Figure()
        
        fig_sak436.add_trace(
            go.Scatter(x=sak_1['Date/Time'], y=sak_1['SAK 436'], name='sample 1', line=dict(color=colors['marker']))
            )
        fig_sak436.add_trace(
            go.Scatter(x=sak_2['Date/Time'], y=sak_2['SAK 436'], name='sample 2', line=dict(color='cyan'))
            )
        fig_sak436.add_trace(
            go.Scatter(x=sak_3['Date/Time'], y=sak_3['SAK 436'], name='sample 3', line=dict(color='yellow'))
            )
        
        # Sample 1 lower Limit
        fig_sak436.add_trace(
            go.Scatter(x=list(sak_1['Date/Time']), y = [sak436_s1_base_lower_limit[0]] * len(sak_1.index), 
                       name='Upper / Lower \nLimit Sample 1', mode='lines',
                       visible=False, line=dict(color=colors['marker'], dash='dash'))
            )
        
        # Sample 1 Upper Limit
        fig_sak436.add_trace(
            go.Scatter(x=list(sak_1['Date/Time']), y = [sak436_s1_base_upper_limit[0]] * len(sak_1.index), 
                       mode='lines', showlegend=False,
                       visible=False, line=dict(color=colors['marker'], dash='dash'))
            )
        
        # Sample 2 lower Limit
        fig_sak436.add_trace(
            go.Scatter(x=list(sak_2['Date/Time']), y = [sak436_s2_base_lower_limit[0]] * len(sak_2.index), 
                       name='Upper / Lower \nLimit Sample 2', mode='lines',
                       visible=False, line=dict(color='cyan', dash='dash'))
            )
        
        # Sample 2 Upper Limit
        fig_sak436.add_trace(
            go.Scatter(x=list(sak_2['Date/Time']), y = [sak436_s2_base_upper_limit[0]] * len(sak_2.index), 
                       mode='lines', showlegend=False,
                       visible=False, line=dict(color='cyan', dash='dash'))
            )
        
        # Sample 3 lower Limit
        fig_sak436.add_trace(
            go.Scatter(x=list(sak_3['Date/Time']), y = [sak436_s3_base_lower_limit[0]] * len(sak_3.index), 
                       name='Upper / Lower \nLimit Sample 3', mode='lines',
                       visible=False, line=dict(color='yellow', dash='dash'))
            )
        
        # Sample 3 Upper Limit
        fig_sak436.add_trace(
            go.Scatter(x=list(sak_3['Date/Time']), y = [sak436_s3_base_upper_limit[0]] * len(sak_3.index), 
                       mode='lines', showlegend=False,
                       visible=False, line=dict(color='yellow', dash='dash'))
            )
        
        #fig_sak436.show(renderer='png')
        
        # sac436 general settings
        fig_sak436.update_layout(
            updatemenus=[
                dict(
                    active=0,
                    buttons=list([
                        dict(label='all Samples',
                             method='update',
                             args=[{'visible': [True, True, True, False, False, False, False, False, False]},
                                   {'title': 'all samples'}]),
                        dict(label='sample 1',
                             method='update',
                             args=[{'visible': [True, False, False, True, True, False, False, False, False]},
                                   {'title': 'sample 1'}]),
                        dict(label='sample 2',
                             method='update',
                             args=[{'visible': [False, True, False, False, False, True, True, False, False]},
                                   {'title': 'sample 2'}]),
                        dict(label='sample 3',
                             method='update',
                             args=[{'visible': [False, False, True, False, False, False, False, True, True]},
                                   {'title': 'sample 3'}])
                        
                        ])
                    )
                ],
            showlegend=True
            )
                             
        fig_sak436.update_layout(
            plot_bgcolor=colors['background'],
            xaxis=dict(title='date/time', 
                       showgrid=False, 
                       linewidth=2, 
                       #linecolor=colors['text'],
                       zeroline=False
                       ),
            yaxis=dict(title='spectral absorbance in 1/m',
                       # dtick=0.02,
                       # range=[0.1, 0.30], 
                       # 
                       showgrid=False, 
                       linewidth=2, 
                       #linecolor=colors['text'],
                       zeroline=False
                       ),
            paper_bgcolor=colors['background'],
            font_color=colors['text']
        )
        
        #fig_sak436.show(renderer='png')
        
# =============================================================================
#         # read and safe flow cytometry data (fcm)
# =============================================================================       
        # read scores fcm file containing all data ('fcm_scores')
        fcm_scores_file = folder_fcm+'fcm_scores.xlsx'        
        fcm_scores = pd.read_excel(fcm_scores_file)
        
        # read the continuously updated file 'fcm_results' 
        fcm_file = folder_fcm+'fcm_results.xlsx'
        fcm_new = pd.read_excel(fcm_file)
        fcm_new['Date/Time'] = pd.to_datetime(fcm_new['Date'] + '' + fcm_new['Time'], format='%Y-%m-%d%H:%M:%S')
        
        # check if there is any data in 'fcm_results'
        if fcm_new.empty:
            print('Waiting for first FCM sample')      
        else:
            fcm_old_index = fcm_scores.iloc[-1,2]
            fcm_new_index = fcm_new.iloc[-1,0]
            
            # check if there is a new reading in 'fcm_results' compared to the latest one in 'fcm_scores' 
            if fcm_old_index == fcm_new_index:
                print('Waiting for new FCM Sample')      
            else: 
                print('New FCM Sample')
                fcm_added = fcm_new.iloc[-1:]
                fcm_scores = pd.concat([fcm_scores, fcm_added], ignore_index=True)
                fcm_scores.to_excel(fcm_scores_file, index=False)
        
# =============================================================================
#         # plot fcm data (dashboard update)
# =============================================================================
        fcm_added = fcm_scores.iloc[1:]
        fcm_added_1 = fcm_added[fcm_added['Sample ID'] == 1]
        fcm_added_2 = fcm_added[fcm_added['Sample ID'] == 2]
        fcm_added_3 = fcm_added[fcm_added['Sample ID'] == 3]
        
        # Sample 1 mean values
        tcc_mean_s1 = fcm_added_1.iloc[0:19, 6:7].mean()
        hna_a_mean_s1 = fcm_added_1.iloc[0:19, 7:8].mean()
        
        # Sample 1 Std values
        tcc_std_s1 = fcm_added_1.iloc[0:19, 6:7].std()
        hna_a_std_s1 = fcm_added_1.iloc[0:19, 7:8].std()
        
        # calculate mean and STD. Threshold factor from literature 3 (10.1016/j.watres.2021.117387)
        t = 3
        
        # sample 1 Define upper and lower limits for respective baselines
        tcc_s1_base_upper_limit = (tcc_mean_s1 + t * tcc_std_s1)
        tcc_s1_base_lower_limit = (tcc_mean_s1 - t * tcc_std_s1)
        hna_a_s1_base_upper_limit = (hna_a_mean_s1 + t * hna_a_std_s1)
        hna_a_s1_base_lower_limit = (hna_a_mean_s1 - t * hna_a_std_s1)
        
        # Sample 2 mean values
        tcc_mean_s2 = fcm_added_2.iloc[0:19, 6:7].mean()
        hna_a_mean_s2 = fcm_added_2.iloc[0:19, 7:8].mean()
        
        # Sample 2 Std values
        tcc_std_s2 = fcm_added_2.iloc[0:19, 6:7].std()
        hna_a_std_s2 = fcm_added_2.iloc[0:19, 7:8].std()
        
        # sample 2 Define upper and lower limits for respective baselines
        tcc_s2_base_upper_limit = (tcc_mean_s2 + t * tcc_std_s2)
        tcc_s2_base_lower_limit = (tcc_mean_s2 - t * tcc_std_s2)
        hna_a_s2_base_upper_limit = (hna_a_mean_s2 + t * hna_a_std_s2)
        hna_a_s2_base_lower_limit = (hna_a_mean_s2 - t * hna_a_std_s2)
        
        # Sample 3 mean values
        tcc_mean_s3 = fcm_added_3.iloc[0:19, 6:7].mean()
        hna_a_mean_s3 = fcm_added_3.iloc[0:19, 7:8].mean()
        
        # Sample 3 Std values
        tcc_std_s3 = fcm_added_3.iloc[0:19, 6:7].std()
        hna_a_std_s3 = fcm_added_3.iloc[0:19, 7:8].std()
        
        # sample 3 Define upper and lower limits for respective baselines
        tcc_s3_base_upper_limit = (tcc_mean_s3 + t * tcc_std_s3)
        tcc_s3_base_lower_limit = (tcc_mean_s3 - t * tcc_std_s3)
        hna_a_s3_base_upper_limit = (hna_a_mean_s3 + t * hna_a_std_s3)
        hna_a_s3_base_lower_limit = (hna_a_mean_s3 - t * hna_a_std_s3)
        
# =============================================================================
#         # TCC
# =============================================================================
        fig_tcc = go.Figure()
        
        fig_tcc.add_trace(
            go.Scatter(x=fcm_added_1['Date/Time'], y=fcm_added_1['TCC'], name='sample 1', line=dict(color=colors['marker']))
            )
        fig_tcc.add_trace(
            go.Scatter(x=fcm_added_2['Date/Time'], y=fcm_added_2['TCC'], name='sample 2', line=dict(color='cyan'))
            )
        fig_tcc.add_trace(
            go.Scatter(x=fcm_added_3['Date/Time'], y=fcm_added_3['TCC'], name='sample 3', line=dict(color='yellow'))
            )
        
        # Sample 1 lower Limit
        fig_tcc.add_trace(
            go.Scatter(x=list(fcm_added_1['Date/Time']), y = [tcc_s1_base_lower_limit[0]] * len(fcm_added_1.index), 
                       name='upper / lower \nlimit sample 1', mode='lines',
                       visible=False, line=dict(color=colors['marker'], dash='dash'))
            )
        
        # Sample 1 Upper Limit
        fig_tcc.add_trace(
            go.Scatter(x=list(fcm_added_1['Date/Time']), y = [tcc_s1_base_upper_limit[0]] * len(fcm_added_1.index), 
                       mode='lines', showlegend=False,
                       visible=False, line=dict(color=colors['marker'], dash='dash'))
            )
        
        # Sample 2 lower Limit
        fig_tcc.add_trace(
            go.Scatter(x=list(fcm_added_2['Date/Time']), y = [tcc_s2_base_lower_limit[0]] * len(fcm_added_2.index), 
                       name='upper / lower \nlimit sample 2', mode='lines',
                       visible=False, line=dict(color='cyan', dash='dash'))
            )
        
        # Sample 2 Upper Limit
        fig_tcc.add_trace(
            go.Scatter(x=list(fcm_added_2['Date/Time']), y = [tcc_s2_base_upper_limit[0]] * len(fcm_added_2.index), 
                       mode='lines', showlegend=False,
                       visible=False, line=dict(color='cyan', dash='dash'))
            )
        
        # Sample 3 lower Limit
        fig_tcc.add_trace(
            go.Scatter(x=list(fcm_added_3['Date/Time']), y = [tcc_s3_base_lower_limit[0]] * len(fcm_added_3.index), 
                       name='upper / lower \nlimit sample 3', mode='lines',
                       visible=False, line=dict(color='yellow', dash='dash'))
            )
        
        # Sample 3 Upper Limit
        fig_tcc.add_trace(
            go.Scatter(x=list(fcm_added_3['Date/Time']), y = [tcc_s3_base_upper_limit[0]] * len(fcm_added_3.index), 
                       mode='lines', showlegend=False,
                       visible=False, line=dict(color='yellow', dash='dash'))
            )
        
        #fig_tcc.show(renderer='png')
        
        # TCC general settings
        fig_tcc.update_layout(
            updatemenus=[
                dict(
                    active=0,
                    buttons=list([
                        dict(label='all samples',
                             method='update',
                             args=[{'visible': [True, True, True, False, False, False, False, False, False]},
                                   {'title': 'total cell count (TCC)'}]),
                        dict(label='sample 1',
                             method='update',
                             args=[{'visible': [True, False, False, True, True, False, False, False, False]},
                                   {'title': 'total cell count (TCC)'}]),
                        dict(label='sample 2',
                             method='update',
                             args=[{'visible': [False, True, False, False, False, True, True, False, False]},
                                   {'title': 'total cell count (TCC)'}]),
                        dict(label='sample 3',
                             method='update',
                             args=[{'visible': [False, False, True, False, False, False, False, True, True]},
                                   {'title': 'total cell count (TCC)'}])
                        
                        ])
                    )
                ],
            showlegend=True
            )
        
        fig_tcc.update_layout(
            title = 'total cell count (TCC)',
            plot_bgcolor=colors['background'],
            xaxis=dict(title='date/time', 
                       showgrid=False, 
                       linewidth=2, 
                       linecolor=colors['text'],
                       zeroline=False
                       ),
            yaxis=dict(title='cells/mL',
                       # dtick=0.02,
                       # range=[0.1, 0.30], 
                       # 
                       showgrid=False, 
                       linewidth=2, 
                       linecolor=colors['text'],
                       zeroline=False
                       ),
            paper_bgcolor=colors['background'],
            font_color=colors['text']
        ) 
        
        #fig_tcc.show(renderer='png')
        
# =============================================================================
#         # HNA in %
# =============================================================================
        fig_hna_a = go.Figure()
        
        fig_hna_a.add_trace(
            go.Scatter(x=fcm_added_1['Date/Time'], y=fcm_added_1['HNA in %'], name='sample 1', line=dict(color=colors['marker']))
            )
        fig_hna_a.add_trace(
            go.Scatter(x=fcm_added_2['Date/Time'], y=fcm_added_2['HNA in %'], name='sample 2', line=dict(color='cyan'))
            )
        fig_hna_a.add_trace(
            go.Scatter(x=fcm_added_3['Date/Time'], y=fcm_added_3['HNA in %'], name='sample 3', line=dict(color='yellow'))
            )
        
        # Sample 1 lower Limit
        fig_hna_a.add_trace(
            go.Scatter(x=list(fcm_added_1['Date/Time']), y = [hna_a_s1_base_lower_limit[0]] * len(fcm_added_1.index), 
                       name='upper / lower \nlimit sample 1', mode='lines',
                       visible=False, line=dict(color=colors['marker'], dash='dash'))
            )
        
        # Sample 1 Upper Limit
        fig_hna_a.add_trace(
            go.Scatter(x=list(fcm_added_1['Date/Time']), y = [hna_a_s1_base_upper_limit[0]] * len(fcm_added_1.index), 
                       mode='lines', showlegend=False,
                       visible=False, line=dict(color=colors['marker'], dash='dash'))
            )
        
        # Sample 2 lower Limit
        fig_hna_a.add_trace(
            go.Scatter(x=list(fcm_added_2['Date/Time']), y = [hna_a_s2_base_lower_limit[0]] * len(fcm_added_2.index), 
                       name='upper / lower \nlimit sample 2', mode='lines',
                       visible=False, line=dict(color='cyan', dash='dash'))
            )
        
        # Sample 2 Upper Limit
        fig_hna_a.add_trace(
            go.Scatter(x=list(fcm_added_2['Date/Time']), y = [hna_a_s2_base_upper_limit[0]] * len(fcm_added_2.index), 
                       mode='lines', showlegend=False,
                       visible=False, line=dict(color='cyan', dash='dash'))
            )
        
        # Sample 3 lower Limit
        fig_hna_a.add_trace(
            go.Scatter(x=list(fcm_added_3['Date/Time']), y = [hna_a_s3_base_lower_limit[0]] * len(fcm_added_3.index), 
                       name='upper / lower \nlimit sample 3', mode='lines',
                       visible=False, line=dict(color='yellow', dash='dash'))
            )
        
        # Sample 3 Upper Limit
        fig_hna_a.add_trace(
            go.Scatter(x=list(fcm_added_3['Date/Time']), y = [hna_a_s3_base_upper_limit[0]] * len(fcm_added_3.index), 
                       mode='lines', showlegend=False,
                       visible=False, line=dict(color='yellow', dash='dash'))
            )
        
        #fig_hna_a.show(renderer='png')
        
        # HNA general settings
        fig_hna_a.update_layout(
            updatemenus=[
                dict(
                    active=0,
                    buttons=list([
                        dict(label='all samples',
                             method='update',
                             args=[{'visible': [True, True, True, False, False, False, False, False, False]},
                                   {'title': 'high nucleic acid (HNA) amount in %'}]),
                        dict(label='sample 1',
                             method='update',
                             args=[{'visible': [True, False, False, True, True, False, False, False, False]},
                                   {'title': 'high nucleic acid (HNA) amount in %'}]),
                        dict(label='sample 2',
                             method='update',
                             args=[{'visible': [False, True, False, False, False, True, True, False, False]},
                                   {'title': 'high nucleic acid (HNA) amount in %'}]),
                        dict(label='sample 3',
                             method='update',
                             args=[{'visible': [False, False, True, False, False, False, False, True, True]},
                                   {'title': 'high nucleic acid (HNA) amount in %'}])
                        
                        ])
                    )
                ],
            showlegend=True
            )        
        
        fig_hna_a.update_layout(
            title='high nucleic acid (HNA) amount in %',
            plot_bgcolor=colors['background'],
            xaxis=dict(title='date/time', 
                       showgrid=False, 
                       linewidth=2, 
                       linecolor=colors['text'],
                       zeroline=False
                       ),
            yaxis=dict(title='%',
                       # dtick=0.02,
                       # range=[0.1, 0.30], 
                       # 
                       showgrid=False, 
                       linewidth=2, 
                       linecolor=colors['text'],
                       zeroline=False
                       ),
            paper_bgcolor=colors['background'],
            font_color=colors['text']
        )  
        
        #fig_hna_a.show(renderer='png')
               
# =============================================================================
#         # if you want to include the current fingerprint (density plot) to the dashboard, not veryfied yet
#         image_folder = 'C:/Users/Public/flow_cytometry'
#         png_files = glob.glob(os.path.join(image_folder, '*.png'))
#         newest_png = max(png_files, key=os.path.getctime)
# =============================================================================
        
        
# =============================================================================
#     # General Dashboard settings and allignment
# =============================================================================
    return [   
        html.Div([
            html.Div(children=[
                html.H1(
                    children='drinking water quality monitoring', 
                    style={
                        'textAlign': 'center',
                        'color': colors['text']
                    }
                    ),
                html.H2(
                    children='an approach for detailed flow cytometry and fluorescence spectroscopy analysis',
                    style={
                        'textAlign': 'center', 'color': colors['text']}),
                html.H2(children='flow cytometry online data',
                        style={'textAlign': 'left', 'color':colors['text']}
                    )
                ])
            ]),
# =============================================================================
#         html.Div([
#             html.Img(id='png1', src=newest_png, title=newest_png, style={'max-width': '100%'})
#             ]),
# =============================================================================
        html.Div([
            html.Div([
                dcc.Graph(id='graph1', figure=fig_tcc)],
                    className="six columns"),
            html.Div([
                dcc.Graph(id='graph2', figure=fig_hna_a)],
                    className='six columns'),
                ], className='row'),
            
# =============================================================================
#             html.Div([
#                 html.H3(children='TCC Boxplot', 
#                         style={'textAlign': 'center', 'color':colors['text']})],
#                 className="three columns"),
# =============================================================================
        html.Div([
            html.Div([
                html.H2(children='fluorescence spectroscopy fingerprint', 
                        style={'textAlign': 'left', 'color':colors['text']}),
                dcc.Graph(id='graph10', figure=fp_c1)], 
                className="three columns"),
            html.Div([
                html.H2(children='fluorescence spectroscopy online data', 
                        style={'textAlign': 'left', 'color':colors['text']}),
                dcc.Graph(id='graph12', figure=fig_c1)],
                className="eight columns"),
# =============================================================================
#             html.Div([
#                 html.H3(children='Boxplot', 
#                         style={'textAlign': 'center', 'color':colors['text']}),
#                 dcc.Graph(id='graph12', figure=boxplot1)],
#                 className="three columns"),
# =============================================================================
        ], className="row"),   
        # html.Hr(),
        html.Div([
            html.Div([
                dcc.Graph(id='graph20', figure=fp_c2)], 
                className="three columns"),
            html.Div([
                dcc.Graph(id='graph22', figure=fig_c2)], 
                className="eight columns"),
# =============================================================================
#             html.Div([
#                 dcc.Graph(id='graph23', figure=boxplot2)],
#                 className="three columns"),
# =============================================================================
        ], className="row"),       
        html.Div([
            html.Div([
                dcc.Graph(id='graph30', figure=fp_c3)], 
                className="three columns"),
            html.Div([
                dcc.Graph(id='graph32', figure=fig_c3)], 
                className="eight columns"),
# =============================================================================
#             html.Div([
#                 dcc.Graph(id='graph33', figure=boxplot3)],
#                 className="three columns"),
# =============================================================================
        ], className="row"),   
        html.Div([
            html.Div([
                dcc.Graph(id='graph40', figure=fp_c4)], 
                className="three columns"),
            html.Div([
                dcc.Graph(id='graph42', figure=fig_c4)], 
                className="eight columns"),
# =============================================================================
#             html.Div([
#                 dcc.Graph(id='graph43', figure=boxplot4)],
#                 className="three columns"),
# =============================================================================
        ], className="row"),   
        html.Div([
            html.Div([
                dcc.Graph(id='graph50', figure=fp_c5)], 
                className="three columns"),
            html.Div([
                dcc.Graph(id='graph52', figure=fig_c5)], 
                className="eight columns"),
# =============================================================================
#             html.Div([
#                 dcc.Graph(id='graph53', figure=boxplot5)],
#                 className="three columns"),
# =============================================================================
        ], className="row"),   
        html.Div([
            html.Div([
                dcc.Graph(id='graph60', figure=fp_c6)], 
                className="three columns"),
            html.Div([
                dcc.Graph(id='graph62', figure=fig_c6)], 
                className="eight columns"),
# =============================================================================
#             html.Div([
#                 dcc.Graph(id='graph63', figure=boxplot6)],
#                 className="three columns"),
# =============================================================================
        ], className="row"),
        html.Div([
            html.Div([
                html.H2(children='spectral absorbance coefficient at 254 nm', 
                        style={'textAlign': 'left', 'color':colors['text']}),
                dcc.Graph(id='graph12', figure=fig_sak254)],
                className="six columns"),
            html.Div([
                html.H2(children='spectral absorbance coefficient at 436 nm',
                        style={'textAlign': 'left', 'color':colors['text']}),
                dcc.Graph(id='graph13', figure=fig_sak436)],
                className='six columns'),
        ], className='row')
    ]

        

if __name__ == '__main__':
    app.run_server(debug=True)
    
print('\nVisit http://127.0.0.1:8050/ in your web browser.')
    
    
